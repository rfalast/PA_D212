import logging
import openpyxl

from os.path import exists
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font, Side, Border, PatternFill
from openpyxl.worksheet.cell_range import CellRange
from pandas import DataFrame
from model.BaseModel import BaseModel
from model.analysis.DatasetAnalyzer import DatasetAnalyzer
from model.analysis.PCA_Analysis import PCA_Analysis
from model.analysis.PlotGenerator import PlotGenerator
from model.analysis.StatisticsGenerator import StatisticsGenerator
from model.constants.BasicConstants import ANALYSIS_TYPE, ANALYZE_DATASET_FULL, ANALYZE_DATASET_INITIAL, MT_OPTIONS, \
    MT_LOGISTIC_REGRESSION, MT_LINEAR_REGRESSION, MT_KNN_CLASSIFICATION, MT_RF_REGRESSION
from model.constants.DatasetConstants import COLUMN_KEY, COLUMN_NA_COUNT, UNIQUE_COLUMN_FLAG, INT64_COLUMN_KEY, \
    COLUMN_COUNT_KEY, FLOAT64_COLUMN_KEY, UNIQUE_COLUMN_VALUES, OBJECT_COLUMN_KEY, COLUMN_TOTAL_COUNT_KEY, \
    UNIQUE_COLUMN_RATIOS, OBJECT_COLUMN_COUNT_KEY, BOOL_COLUMN_KEY, BOOL_COLUMN_COUNT_KEY, UNIQUE_COLUMN_LIST_KEY, \
    YES_NO_LIST
from model.constants.ModelConstants import LM_P_VALUE, LM_FEATURE_NUM, LM_COEFFICIENT, LM_STANDARD_ERROR, \
    LM_T_STATISTIC, LM_INITIAL_MODEL, LM_FINAL_MODEL, LM_VIF, LM_LS_CONF_INT, LM_RS_CONF_INT, \
    LM_LAGRANGE_MULTIPLIER_STATISTIC, LM_JARQUE_BERA_STATISTIC, LM_JARQUE_BERA_PROB, KNN_ACCURACY_SCORE, LM_VALID_TYPES
from model.constants.ReportConstants import REPORT_INITIAL, REPORT_BASE_NAME, REPORT_DF_BASE_NAME, BASIC_ITEM_HEADER, \
    BASIC_COLUMN_HEADER, BASIC_TYPE_HEADER, BASIC_COUNT_HEADER, NA_COLUMN_LABEL, FORMAT_NUMBER_COMMA, MEAN_HEADER, \
    MEDIAN_HEADER, VARIANCE_HEADER, STD_DEV_HEADER, SKEW_HEADER, MIN_HEADER, MAX_HEADER, VALUES_ACTUAL, TRUE_COUNT, \
    FALSE_COUNT, DISTINCT_VALUE_COUNT, VALUE_UNIQUE_HEADER, VALUES_TOO_MANY, BORDER_THICK, OBJ_DATA_FIELD_COLUMN_VAL, \
    OBJ_DATA_FIELD_COUNT, OBJ_DATA_FIELD_COLUMN_PCT, OBJ_DATA_FIELD_CUM_PCT, FORMAT_PERCENTAGE, \
    DEFAULT_ROW_HEADER_SPACER, DEFAULT_ROW_SPACER, MINIMUM_OFFSET, REPORT_ORIG_DF_TAB_NAME, REPORT_CLEAN_DF_TAB_NAME, \
    REPORT_NORM_DF_TAB_NAME, REPORT_PAC_TAB_NAME, BASIC_SINGLE_SPACE, COLUMN_HEADER, CORRELATION_HEADER, \
    CHI_SQUARED_HEADER, CORRELATION_LEVEL, DISTRIBUTION, VALUES_UNIQUE, TARGET_HEADER, NUMBER_OF_FEATURES_HEADER, \
    CONSTANT_HEADER, R_SQUARED_HEADER, ADJ_R_SQUARED_HEADER, AIC_HEADER, BIC_HEADER, FEATURE_NUMBER, \
    FEATURE_NAME_HEADER, P_VALUE_HEADER, COEFFICIENT_HEADER, F_STATISTIC_HEADER, STANDARD_ERROR_HEADER, \
    T_STATISTIC_HEADER, P_VALUE_F_STATISTIC_HEADER, VIF_HEADER, DURBAN_WATSON_STATISTIC, \
    NUMBER_OF_OBS, LOG_LIKELIHOOD, DEGREES_OF_FREEDOM_MODEL, DEGREES_OF_FREEDOM_RESID, RESIDUAL_STD_ERROR, \
    BREUSCH_PAGAN_P_VALUE, MODEL_ACCURACY, MODEL_PRECISION, MODEL_RECALL, MODEL_F1_SCORE, JARQUE_BERA_STATISTIC, \
    PSEUDO_R_SQUARED_HEADER, AIC_SCORE, BIC_SCORE, MODEL_CONSTANT, MODEL_AVG_PRECISION, MODEL_ROC_SCORE, \
    MODEL_BEST_SCORE, MODEL_BEST_PARAMS, MODEL_MEAN_ABSOLUTE_ERROR, MODEL_MEAN_SQUARED_ERROR, \
    MODEL_ROOT_MEAN_SQUARED_ERROR
from model.constants.StatisticsConstants import ALL_CORRELATIONS, DIST_NAME
from util.CommonUtils import get_tuples_from_list_with_specific_field
from util.ExcelManager import ExcelManager
from model.constants.PlotConstants import PLOT_TYPE_HIST, PLOT_TYPE_BOX_PLOT, PLOT_TYPE_BAR_CHART, BASE_OUTPUT_PATH, \
    PLOT_TYPE_HEATMAP, GENERAL, PLOT_TYPE_SCATTER_CHART, PLOT_TYPE_BIVARIATE_COUNT, PLOT_TYPE_JOINT_PLOT, \
    PLOT_TYPE_Q_Q_PLOT, PLOT_TYPE_STD_RESIDUAL, PLOT_TYPE_BIVARIATE_BOX, PLOT_TYPE_Q_Q_RESIDUAL_PLOT, \
    PLOT_TYPE_LONG_ODDS, PLOT_TYPE_CM_HEATMAP, PLOT_TYPE_ROC_AUC


# class for generating output file
class ReportGenerator(BaseModel):
    # init method
    def __init__(self, output_path_override=None):
        # call superclass
        super().__init__()

        # initialize logger
        self.logger = logging.getLogger(__name__)

        # log that we've been called.
        self.logger.debug("A request to instantiate a ReportGenerator has been made.")

        # check if the override variable was passed
        if self.is_valid(output_path_override, str):
            # log that we're overriding the path
            self.logger.debug(f"overriding output path to [{output_path_override}]")

            # generate the output paths
            self.initial_output_path = output_path_override + REPORT_INITIAL + REPORT_BASE_NAME
            self.output_path = output_path_override + REPORT_BASE_NAME
            self.dataframe_path = output_path_override + REPORT_DF_BASE_NAME
            self.report_path = output_path_override

            # validate that the overriden output path is valid
            if not exists(output_path_override):
                raise SyntaxError(f"The overriden path does not exist, or is not valid.")
        # generate the full path
        else:
            # log that we're not overriding the path
            self.logger.debug("Not overriding output path.")

            # generate the output paths including file
            self.initial_output_path = BASE_OUTPUT_PATH + REPORT_INITIAL + REPORT_BASE_NAME
            self.output_path = BASE_OUTPUT_PATH + REPORT_BASE_NAME
            self.dataframe_path = BASE_OUTPUT_PATH + REPORT_DF_BASE_NAME
            self.report_path = BASE_OUTPUT_PATH

        # define internal variables
        self.the_workbook = None
        self.data_analyzer = None
        self.plot_generator = None
        self.stat_generator = None
        self.initial_model = None
        self.final_model = None

        # log the output path
        self.logger.info("final report output path set to [{self.output_path}]")

    # autosize a worksheet
    @staticmethod
    def autosize_worksheet(the_ws):
        dims = {}
        for row in the_ws.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))
        for col, value in dims.items():
            the_ws.column_dimensions[col].width = value

    # generate an Excel report
    def generate_excel_report(self, the_dataset_analyzer: DatasetAnalyzer, the_plot_generator: PlotGenerator,
                              stat_generator: StatisticsGenerator, the_model_type: str, the_type=None):
        # log that we've been invoked
        self.logger.debug(f"Generating an Excel report for [{the_model_type}][{the_type}]")

        # run validations
        # validate that the output path is not None
        if self.output_path is None:
            raise IOError("The output path is None.")
        elif not self.is_valid(the_dataset_analyzer, DatasetAnalyzer):
            raise SyntaxError("The DatasetAnalyzer argument was None or incorrect type.")
        elif not self.is_valid(the_plot_generator, PlotGenerator):
            raise SyntaxError("The PlotGenerator argument was None or incorrect type.")
        elif not self.is_valid(stat_generator, StatisticsGenerator) and the_type == ANALYZE_DATASET_FULL:
            raise SyntaxError("The StatisticsGenerator argument was None or incorrect type.")
        elif not self.is_valid(the_model_type, str) and the_type == ANALYZE_DATASET_FULL:
            raise SyntaxError("The the_model_type argument was None or incorrect type.")
        elif the_model_type not in MT_OPTIONS:
            raise AttributeError("The the_model_type argument was None or incorrect type.")

        # store the analyzer and plot generator
        self.data_analyzer = the_dataset_analyzer
        self.plot_generator = the_plot_generator
        self.stat_generator = stat_generator

        # check if this is a full analysis run
        if the_type == ANALYZE_DATASET_FULL:
            # if this is a full run, we need to retrieve the initial model.
            self.initial_model = the_dataset_analyzer.linear_model_storage[LM_INITIAL_MODEL]

            if the_model_type in [MT_LOGISTIC_REGRESSION, MT_LINEAR_REGRESSION]:
                # retrieve the final model if we have Linear or Logistic model
                self.final_model = the_dataset_analyzer.linear_model_storage[LM_FINAL_MODEL]

        # create workbook instance
        self.the_workbook = openpyxl.Workbook()

        # log that we're creating basic data tab
        self.logger.debug("Creating DATASET BASICS tab.")

        # create basic analysis section
        self.__create_basic_analysis_sheet__(0)

        # create integer analysis section
        self.__create_integer_analysis_sheet__(1, the_type=the_type)

        # create integer field analysis section
        self.__create_int_field_analysis_sheet__(2, the_type=the_type)

        # create float analysis section
        self.__create_float_analysis_sheet__(3, the_type=the_type)

        # create float field analysis section
        self.__create_float_field_analysis_sheet__(4, the_type=the_type)

        # create object analysis section
        self.__create_object_analysis_sheet__(5)

        # create counts of each distinct object value
        self.__create_object_count_sheet__(6, the_type=the_type)

        # create boolean analysis section
        self.__create_boolean_analysis_sheet__(7)

        # create counts of each distinct boolean value
        self.__create_boolean_count_sheet__(8, the_type=the_type)

        # check if this is an INITIAL or FULL report.  If it is INITIAL, we skip these tabs.
        if the_type == ANALYZE_DATASET_FULL:
            # create the correlations tab
            self.__create_correlations_sheet__(the_model_type=the_model_type, the_index=9)

            # create the chi-squared tab
            self.__create_chi_squared_sheet__(the_index=10)

            # create the FINAL model tab if LOGISTIC or LINEAR model
            if the_model_type in [MT_LOGISTIC_REGRESSION, MT_LINEAR_REGRESSION]:
                # create the INITIAL model tab
                self.__create_model_results_sheet__(the_index=11,
                                                    the_type=ANALYZE_DATASET_INITIAL,
                                                    the_model_type=the_model_type)

                # create the FINAL model tab
                self.__create_model_results_sheet__(the_index=12,
                                                    the_type=ANALYZE_DATASET_FULL,
                                                    the_model_type=the_model_type)

            elif the_model_type in [MT_KNN_CLASSIFICATION, MT_RF_REGRESSION]:
                # should create only a single tab.
                self.__create_model_results_sheet__(the_index=11,
                                                    the_type=ANALYZE_DATASET_FULL,
                                                    the_model_type=the_model_type)

        # log that we're about to save the workbook
        self.logger.debug(f"About to save workbook to [{self.output_path}]")

        # save the FINAL report. the_type is ANALYSIS_TYPE[0] or FULL
        if the_type is ANALYZE_DATASET_FULL:
            # log what we're saving
            self.logger.debug(f"About to save [{self.output_path}]")

            # save the workbook
            self.the_workbook.save(self.output_path)
        # save the INITIAL report. the_type is ANALYSIS_TYPE[1] or INITIAL
        elif the_type is ANALYSIS_TYPE[1]:
            # log what we're saving
            self.logger.debug(f"About to save [{self.initial_output_path}]")

            # save the workbook
            self.the_workbook.save(self.initial_output_path)
        # not sure how this happens
        else:
            raise RuntimeError("unexpected output type.")

    # private method to create the basic analysis sheet
    def __create_basic_analysis_sheet__(self, the_index):
        # log that we've been called
        self.logger.info("A request to generate the basic data tab has been made.")

        # create the basic analysis tab
        self.the_workbook.create_sheet(index=the_index, title="DATA SET BASICS")

        # get the sheet
        the_ws = self.the_workbook.worksheets[the_index]

        # get the dictionaries from the data_analyzer storage
        the_columns = self.data_analyzer.storage[COLUMN_KEY]

        # get the dict of NaN count per column
        nan_count_dict = self.data_analyzer.storage[COLUMN_NA_COUNT]

        # get the dict of whether not this column is unique
        unique_column_dict = self.data_analyzer.storage[UNIQUE_COLUMN_LIST_KEY]

        # log that we're about loop over the columns
        self.logger.debug("About to start looping over the columns.")

        # let's select cell A1 to start writing out the headers
        the_ws["A1"].value = BASIC_ITEM_HEADER
        the_ws["A1"].alignment = Alignment(horizontal='center')
        the_ws["A1"].font = Font(bold=True)
        the_ws.column_dimensions['A'].width = 9

        the_ws["B1"].value = BASIC_COLUMN_HEADER
        the_ws["B1"].font = Font(bold=True)
        the_ws.column_dimensions['B'].width = 18

        the_ws["C1"].value = BASIC_TYPE_HEADER
        the_ws["C1"].alignment = Alignment(horizontal='left')
        the_ws["C1"].font = Font(bold=True)
        the_ws.column_dimensions['C'].width = 9

        the_ws["D1"].value = BASIC_COUNT_HEADER
        the_ws["D1"].alignment = Alignment(horizontal='center')
        the_ws["D1"].font = Font(bold=True)
        the_ws.column_dimensions['D'].width = 9

        the_ws["E1"].value = NA_COLUMN_LABEL
        the_ws["E1"].alignment = Alignment(horizontal='center')
        the_ws["E1"].font = Font(bold=True)
        the_ws.column_dimensions['E'].width = 10

        the_ws["F1"].value = VALUES_UNIQUE
        the_ws["F1"].alignment = Alignment(horizontal='center')
        the_ws["F1"].font = Font(bold=True)
        the_ws.column_dimensions['E'].width = 15

        # we need to freeze the top row.  I'm not sure why, but you have to be one row below what you think.
        the_ws.freeze_panes = the_ws["B2"]

        # define counter
        row_count = 1

        # get the columns
        for column in the_columns:
            # log what we have
            self.logger.debug(f"column[{column}] of type[{the_columns[column]}]")

            # write out the information
            the_ws["A1"].offset(row=row_count, column=0).value = row_count
            the_ws["A1"].offset(row=row_count, column=0).alignment = Alignment(horizontal='center')

            # column name
            the_ws["A1"].offset(row=row_count, column=1).value = column

            # column data type
            the_ws["A1"].offset(row=row_count, column=2).value = str(the_columns[column])

            # column count
            the_ws["A1"].offset(row=row_count, column=3).value = self.data_analyzer.the_df[column].count()
            the_ws["A1"].offset(row=row_count, column=3).number_format = FORMAT_NUMBER_COMMA
            the_ws["A1"].offset(row=row_count, column=3).alignment = Alignment(horizontal='center')

            # write out the NaN count
            the_ws["A1"].offset(row=row_count, column=4).value = nan_count_dict[column]
            the_ws["A1"].offset(row=row_count, column=4).alignment = Alignment(horizontal='center')
            the_ws["A1"].offset(row=row_count, column=4).number_format = FORMAT_NUMBER_COMMA

            # write out if unique
            if column in unique_column_dict:
                the_ws["A1"].offset(row=row_count, column=5).value = YES_NO_LIST[0]
            else:
                the_ws["A1"].offset(row=row_count, column=5).value = YES_NO_LIST[1]
            the_ws["A1"].offset(row=row_count, column=5).alignment = Alignment(horizontal='center')

            # increment row count
            row_count = row_count + 1

    # create integer analysis sheet
    def __create_integer_analysis_sheet__(self, the_index, the_type=None):
        # log that we've been called
        self.logger.info("A request to generate the integer data tab has been made.")

        # create the basic analysis tab
        self.the_workbook.create_sheet(index=the_index, title="INTEGER DATA ANALYSIS")

        # get the sheet
        the_ws = self.the_workbook.worksheets[the_index]

        # get the list of integers
        the_list = self.data_analyzer.storage[INT64_COLUMN_KEY]

        # log the number of integers we have
        self.logger.info(f"There list of integers is of size [{len(the_list)}]", )

        # get the dict of what row number
        count_dict = self.data_analyzer.storage[COLUMN_COUNT_KEY]

        # log that we're about to create the headers
        self.logger.debug("About to start creating the header row.")

        # additional variable declaration
        anc_col = "A1"
        row_incr = 0
        col_incr = 0
        dist_dict = {}

        # get the dist_dict from the stat generator if we are ANALYZE_DATASET_FULL
        if the_type == ANALYZE_DATASET_FULL:
            # retrieve the dict of distributions
            dist_dict = self.stat_generator.distribution

        # let's select cell A1 to start writing out the headers
        the_ws["A1"].value = BASIC_ITEM_HEADER
        the_ws["A1"].alignment = Alignment(horizontal='center')
        the_ws["A1"].font = Font(bold=True)
        the_ws.column_dimensions['A'].width = 10

        # NEXT ROW
        col_incr = 1
        the_ws["A1"].offset(row=row_incr, column=col_incr).value = BASIC_COLUMN_HEADER
        the_ws["A1"].offset(row=row_incr, column=col_incr).font = Font(bold=True)
        the_ws.column_dimensions['B'].width = 20

        # NEXT ROW
        col_incr = 2
        the_ws["A1"].offset(row=row_incr, column=col_incr).value = BASIC_COUNT_HEADER
        the_ws["A1"].offset(row=row_incr, column=col_incr).alignment = Alignment(horizontal='center')
        the_ws["A1"].offset(row=row_incr, column=col_incr).font = Font(bold=True)
        the_ws.column_dimensions['C'].width = 9

        # NEXT ROW
        col_incr = 3
        the_ws["A1"].offset(row=row_incr, column=col_incr).value = MEAN_HEADER
        the_ws["A1"].offset(row=row_incr, column=col_incr).alignment = Alignment(horizontal='center')
        the_ws["A1"].offset(row=row_incr, column=col_incr).font = Font(bold=True)
        the_ws.column_dimensions['D'].width = 12

        # NEXT ROW
        col_incr = 4
        the_ws["A1"].offset(row=row_incr, column=col_incr).value = MEDIAN_HEADER
        the_ws["A1"].offset(row=row_incr, column=col_incr).alignment = Alignment(horizontal='center')
        the_ws["A1"].offset(row=row_incr, column=col_incr).font = Font(bold=True)
        the_ws.column_dimensions['E'].width = 12

        # NEXT ROW
        col_incr = 5
        the_ws["A1"].offset(row=row_incr, column=col_incr).value = VARIANCE_HEADER
        the_ws["A1"].offset(row=row_incr, column=col_incr).alignment = Alignment(horizontal='center')
        the_ws["A1"].offset(row=row_incr, column=col_incr).font = Font(bold=True)
        the_ws.column_dimensions['F'].width = 15

        # NEXT ROW
        col_incr = 6
        the_ws["A1"].offset(row=row_incr, column=col_incr).value = STD_DEV_HEADER
        the_ws["A1"].offset(row=row_incr, column=col_incr).alignment = Alignment(horizontal='center')
        the_ws["A1"].offset(row=row_incr, column=col_incr).font = Font(bold=True)
        the_ws.column_dimensions['G'].width = 12

        # NEXT ROW
        col_incr = 7
        the_ws["A1"].offset(row=row_incr, column=col_incr).value = SKEW_HEADER
        the_ws["A1"].offset(row=row_incr, column=col_incr).alignment = Alignment(horizontal='center')
        the_ws["A1"].offset(row=row_incr, column=col_incr).font = Font(bold=True)
        the_ws.column_dimensions['H'].width = 12

        # NEXT ROW
        col_incr = 8
        the_ws["A1"].offset(row=row_incr, column=col_incr).value = MIN_HEADER
        the_ws["A1"].offset(row=row_incr, column=col_incr).alignment = Alignment(horizontal='center')
        the_ws["A1"].offset(row=row_incr, column=col_incr).font = Font(bold=True)
        the_ws.column_dimensions['I'].width = 12

        # NEXT ROW
        col_incr = 9
        the_ws["A1"].offset(row=row_incr, column=col_incr).value = MAX_HEADER
        the_ws["A1"].offset(row=row_incr, column=col_incr).alignment = Alignment(horizontal='center')
        the_ws["A1"].offset(row=row_incr, column=col_incr).font = Font(bold=True)
        the_ws.column_dimensions['J'].width = 12

        # only generate this header if we're doing a FULL
        if the_type == ANALYZE_DATASET_FULL:
            # NEXT ROW
            col_incr = 10
            the_ws["A1"].offset(row=row_incr, column=col_incr).value = DISTRIBUTION
            the_ws["A1"].offset(row=row_incr, column=col_incr).alignment = Alignment(horizontal='center')
            the_ws["A1"].offset(row=row_incr, column=col_incr).font = Font(bold=True)
            the_ws.column_dimensions['K'].width = 16

        # we need to freeze the top row
        the_ws.freeze_panes = the_ws["B2"]

        # log that we're about to start looping over the integers
        self.logger.debug("About to start adding int64 items to page.")

        # define the row_count
        row_count = 1

        # loop over the items in the list
        for next_value in the_list:
            # log the first retrieved item
            self.logger.debug(f"retrieved integer [{next_value}]")

            # write out the item #
            the_ws["A1"].offset(row=row_count, column=0).value = count_dict[next_value]
            the_ws["A1"].offset(row=row_count, column=0).alignment = Alignment(horizontal='center')

            # write out the column name
            the_ws["A1"].offset(row=row_count, column=1).value = next_value
            the_ws["A1"].offset(row=row_count, column=1).alignment = Alignment(horizontal='left')

            # write out the count
            the_ws["A1"].offset(row=row_count, column=2).value = self.data_analyzer.the_df[next_value].count()
            the_ws["A1"].offset(row=row_count, column=2).alignment = Alignment(horizontal='center')
            the_ws["A1"].offset(row=row_count, column=2).number_format = FORMAT_NUMBER_COMMA

            # write out mean
            the_ws["A1"].offset(row=row_count, column=3).value = self.data_analyzer.the_df[next_value].mean()
            the_ws["A1"].offset(row=row_count, column=3).alignment = Alignment(horizontal='center')
            the_ws["A1"].offset(row=row_count, column=3).number_format = "#,##0.000"

            # write out median
            the_ws["A1"].offset(row=row_count, column=4).value = self.data_analyzer.the_df[next_value].median()
            the_ws["A1"].offset(row=row_count, column=4).alignment = Alignment(horizontal='center')
            the_ws["A1"].offset(row=row_count, column=4).number_format = "#,##0.000"

            # write out variance
            the_ws["A1"].offset(row=row_count, column=5).value = self.data_analyzer.the_df[next_value].var()
            the_ws["A1"].offset(row=row_count, column=5).alignment = Alignment(horizontal='center')
            the_ws["A1"].offset(row=row_count, column=5).number_format = "#,##0.000"

            # write out std dev
            the_ws["A1"].offset(row=row_count, column=6).value = self.data_analyzer.the_df[next_value].std()
            the_ws["A1"].offset(row=row_count, column=6).alignment = Alignment(horizontal='center')
            the_ws["A1"].offset(row=row_count, column=6).number_format = "#,##0.0000"

            # write out skew
            the_ws["A1"].offset(row=row_count, column=7).value = self.data_analyzer.the_df[next_value].skew()
            the_ws["A1"].offset(row=row_count, column=7).alignment = Alignment(horizontal='center')
            the_ws["A1"].offset(row=row_count, column=7).number_format = "#,##0.0000"

            # write out min
            the_ws["A1"].offset(row=row_count, column=8).value = self.data_analyzer.the_df[next_value].min()
            the_ws["A1"].offset(row=row_count, column=8).alignment = Alignment(horizontal='center')
            the_ws["A1"].offset(row=row_count, column=8).number_format = FORMAT_NUMBER_COMMA

            # write out max
            the_ws["A1"].offset(row=row_count, column=9).value = self.data_analyzer.the_df[next_value].max()
            the_ws["A1"].offset(row=row_count, column=9).alignment = Alignment(horizontal='center')
            the_ws["A1"].offset(row=row_count, column=9).number_format = FORMAT_NUMBER_COMMA

            if the_type == ANALYZE_DATASET_FULL:
                # write out the distribution
                the_ws["A1"].offset(row=row_count, column=10).value = dist_dict[next_value][DIST_NAME]
                the_ws["A1"].offset(row=row_count, column=10).alignment = Alignment(horizontal='center')

            # increment row_count
            row_count = row_count + 1

    # create integer analysis sheet
    def __create_int_field_analysis_sheet__(self, the_index, the_type=None):
        # log that we've been called
        logging.info("A request to generate the INT field analysis tab has been made.")

        # what I want to do is loop over the list of floats columns, and grab all the images
        # (in this first run, I'm just doing histograms) for the column and place them horizontally
        # onto the generalized rows associated with the float column.  I will construct a general header
        # for the report, where each float column will have a series of rows that display relevant
        # plots for exploring the data.

        # create the basic analysis tab
        self.the_workbook.create_sheet(index=the_index, title="INTEGER DATA FIELD ANALYSIS")

        # get the sheet references
        the_ws = self.the_workbook.worksheets[the_index]

        # Get the list of INT64 columns
        the_list = self.data_analyzer.storage[INT64_COLUMN_KEY]

        # get the dict of what row number the field is in.  Required for the header
        count_dict = self.data_analyzer.storage[COLUMN_COUNT_KEY]

        # get the dict of plots
        plot_dict = self.plot_generator.plot_storage

        # retrieve the list of BOOLEAN columns on key BOOL_COLUMN_KEY
        exclusion_list = self.data_analyzer.storage[BOOL_COLUMN_KEY]

        # log the number of floats we have
        logging.info(f"The list of floats is of size [{len(the_list)}]")

        # set the base cell that we start all calculations
        base_cell_addr = "A1"
        base_col_name_addr = "F1:G1"
        base_col_name_value_addr = "H1:L1"
        the_offset = 0
        relative_offset = 0
        previous_offset = 0
        column_offset = 31  # this is the initial value of the offset

        # define the ranges
        cell_range_F1G1 = CellRange(base_col_name_addr)
        cell_range_H1L1 = CellRange(base_col_name_value_addr)

        # define the image reference.
        the_image = None

        # we now need to loop over the columns in the float list
        for the_column in the_list:
            # log the column that we're getting
            logging.debug(f"column[{the_column}]")

            # Write out the top header in row 1
            # COLUMN A - COLUMN # - HEADER
            the_ws[base_cell_addr].offset(row=the_offset, column=0).value = BASIC_ITEM_HEADER
            the_ws[base_cell_addr].offset(row=the_offset, column=0).alignment = Alignment(horizontal='center')
            the_ws[base_cell_addr].offset(row=the_offset, column=0).font = Font(bold=True)
            the_ws[base_cell_addr].offset(row=the_offset, column=0).border = \
                Border(top=BORDER_THICK, left=BORDER_THICK, right=BORDER_THICK, bottom=BORDER_THICK)
            the_ws.column_dimensions['A'].width = 24

            # COLUMN B - COLUMN # - Actual Value
            the_ws[base_cell_addr].offset(row=the_offset, column=1).value = count_dict[the_column]
            the_ws[base_cell_addr].offset(row=the_offset, column=1).alignment = Alignment(horizontal='center')
            the_ws[base_cell_addr].offset(row=the_offset, column=1).font = Font(bold=True)
            the_ws[base_cell_addr].offset(row=the_offset, column=1).border = \
                Border(top=BORDER_THICK, left=BORDER_THICK, right=BORDER_THICK, bottom=BORDER_THICK)
            the_ws.column_dimensions['B'].width = 9

            # COLUMN F1:G1 - COLUMN NAME - HEADER
            the_ws.merge_cells(start_row=the_offset + 1, start_column=6, end_row=the_offset + 1, end_column=7)
            the_ws[base_cell_addr].offset(row=the_offset, column=5).value = BASIC_COLUMN_HEADER
            the_ws[base_cell_addr].offset(row=the_offset, column=5).alignment = Alignment(horizontal='center')
            the_ws[base_cell_addr].offset(row=the_offset, column=5).font = Font(bold=True)

            # offset the range
            cell_range_F1G1.shift(col_shift=0, row_shift=relative_offset)

            # set the border
            self.set_border(the_ws, cell_range_F1G1)

            # COLUMN H1:L1 - COLUMN NAME - Actual Value
            the_ws.merge_cells(start_row=the_offset + 1, start_column=8, end_row=the_offset + 1, end_column=12)
            the_ws[base_cell_addr].offset(row=the_offset, column=7).value = the_column
            the_ws[base_cell_addr].offset(row=the_offset, column=7).alignment = Alignment(horizontal='left')
            the_ws[base_cell_addr].offset(row=the_offset, column=7).font = Font(bold=True)

            # offset the range
            cell_range_H1L1.shift(col_shift=0, row_shift=relative_offset)

            # set the border
            self.set_border(the_ws, cell_range_H1L1)

            # ***************************************************************************
            #                                HISTOGRAMS
            # ***************************************************************************
            # log the image we're going to import
            logging.debug(f"About to insert image [{plot_dict[the_column][PLOT_TYPE_HIST]}]")

            # create an instance of the image
            the_image = Image(plot_dict[the_column][PLOT_TYPE_HIST])
            the_image.anchor = the_ws[base_cell_addr].offset(row=the_offset + 2, column=1).coordinate

            # log where we're putting the image
            logging.debug(f"Image anchor is [{the_ws[base_cell_addr].offset(row=the_offset + 2, column=1).coordinate}]")

            # add the image to the worksheet
            the_ws.add_image(the_image)

            # ***************************************************************************
            #                                 BOX PLOTS
            # ***************************************************************************

            # log the image we're going to import
            logging.debug(f"About to insert image [{plot_dict[the_column][PLOT_TYPE_BOX_PLOT]}]")

            # create an instance of the image
            the_image = Image(plot_dict[the_column][PLOT_TYPE_BOX_PLOT])
            the_image.anchor = the_ws[base_cell_addr].offset(row=the_offset + 2, column=11).coordinate

            # log where we're putting the image
            logging.debug(f"Image anchor is "
                          f"[{the_ws[base_cell_addr].offset(row=the_offset + 2, column=11).coordinate}]")

            # add the image to the worksheet
            the_ws.add_image(the_image)

            # log all the offset values
            logging.debug(f"prior to adjustment, the_offset[{the_offset}] "
                          f"relative_offset[{relative_offset}] previous_offset[{previous_offset}]")

            # ***************************************************************************
            #                                 QQ PLOTS
            # ***************************************************************************

            # log the image we're going to import
            logging.debug(f"About to insert image [{plot_dict[the_column][PLOT_TYPE_Q_Q_PLOT]}]")

            # create an instance of the image
            the_image = Image(plot_dict[the_column][PLOT_TYPE_Q_Q_PLOT])
            the_image.anchor = the_ws[base_cell_addr].offset(row=the_offset + 2, column=21).coordinate

            # log where we're putting the image
            logging.debug(f"Image anchor is "
                          f"[{the_ws[base_cell_addr].offset(row=the_offset + 2, column=21).coordinate}]")

            # add the image to the worksheet
            the_ws.add_image(the_image)

            # log all the offset values
            logging.debug(f"prior to adjustment, the_offset[{the_offset}] "
                          f"relative_offset[{relative_offset}] previous_offset[{previous_offset}]")

            # ***************************************************************************
            #                                 JOINT PLOTS
            # ***************************************************************************
            if the_type == ANALYZE_DATASET_FULL:
                # get the stat generator
                the_sg = self.stat_generator

                # get the list of tuples for our current column, excluding boolean fields
                tuple_list = the_sg.filter_tuples_by_column(the_sg.all_corr_storage, the_column, exclusion_list)

                # log the tuple_list returned
                self.logger.debug(f"the tuple_list for [{the_column}] is {tuple_list}")

                # loop over all the tuples
                for the_tuple in tuple_list:
                    # log the image we're going to import
                    logging.debug(f"for [{the_column}] inserting image ["
                                  f"{plot_dict[PLOT_TYPE_JOINT_PLOT][the_tuple]}]")

                    # create an instance of the image
                    the_image = Image(plot_dict[PLOT_TYPE_JOINT_PLOT][the_tuple])
                    the_image.anchor = the_ws[base_cell_addr].offset(row=the_offset + 2,
                                                                     column=column_offset).coordinate

                    # log where we're putting the image
                    logging.debug(f"Image anchor is [{the_image.anchor}]")

                    # add the image to the worksheet
                    the_ws.add_image(the_image)

                    # increment the column offset
                    column_offset = column_offset + 10

            # ***************************************************************************
            #                                 END OF PLOTS
            # ***************************************************************************

            # force the offset to MINIMUM_OFFSET
            the_offset = the_offset + MINIMUM_OFFSET + 5

            # capture the relative offset for this loop
            relative_offset = the_offset - previous_offset

            # reset the previous_offset
            previous_offset = the_offset

            # reset the column offset
            column_offset = 31

            # log all the offset values
            logging.debug(f"At end of function, the_offset[{the_offset}] "
                          f"relative_offset[{relative_offset}] previous_offset[{previous_offset}]")

        # log that we are complete.
        logging.debug("Complete with generation of graphs for OBJECT data analysis.")

    # create integer outlier analysis sheet.  Assumption that type is always FULL>
    def __create_int_outlier_analysis_sheet(self, the_index):
        # log that we've been called
        logging.info("A request to generate the INT outlier analysis tab has been made.")

        # create the basic analysis tab
        self.the_workbook.create_sheet(index=the_index, title="INTEGER OUTLIER ANALYSIS")

        # get the sheet references
        the_ws = self.the_workbook.worksheets[the_index]

        # Get the list of INT64 columns
        the_list = self.data_analyzer.storage[INT64_COLUMN_KEY]

    # create float analysis sheet
    def __create_float_analysis_sheet__(self, the_index, the_type=None):
        # log that we've been called
        self.logger.info("A request to generate the float data tab has been made.")

        # create the basic analysis tab
        self.the_workbook.create_sheet(index=the_index, title="FLOAT DATA ANALYSIS")

        # get the sheet
        the_ws = self.the_workbook.worksheets[the_index]

        # get the list of floats
        the_list = self.data_analyzer.storage[FLOAT64_COLUMN_KEY]

        # log the number of floats we have
        self.logger.info(f"There list of floats is of size [{len(the_list)}]")

        # get the dict of what row number
        count_dict = self.data_analyzer.storage[COLUMN_COUNT_KEY]

        # get the dict of NaN count per column
        nan_count_dict = self.data_analyzer.storage[COLUMN_NA_COUNT]

        # log that we're about to create the headers
        self.logger.debug("About to start creating the header row.")

        # additional variable declaration
        anc_col = "A1"
        row_incr = 0
        col_incr = 0
        dist_dict = {}

        # get the dist_dict from the stat generator if we are ANALYZE_DATASET_FULL
        if the_type == ANALYZE_DATASET_FULL:
            # retrieve the dict of distributions
            dist_dict = self.stat_generator.distribution

        # let's select cell A1 to start writing out the headers
        the_ws["A1"].value = BASIC_ITEM_HEADER
        the_ws["A1"].alignment = Alignment(horizontal='center')
        the_ws["A1"].font = Font(bold=True)
        the_ws.column_dimensions['A'].width = 10

        # NEXT COLUMN
        col_incr = 1
        the_ws["A1"].offset(row=row_incr, column=col_incr).value = BASIC_COLUMN_HEADER
        the_ws["A1"].offset(row=row_incr, column=col_incr).font = Font(bold=True)
        the_ws.column_dimensions['B'].width = 20

        # NEXT COLUMN
        col_incr = 2
        the_ws["A1"].offset(row=row_incr, column=col_incr).value = BASIC_COUNT_HEADER
        the_ws["A1"].offset(row=row_incr, column=col_incr).alignment = Alignment(horizontal='center')
        the_ws["A1"].offset(row=row_incr, column=col_incr).font = Font(bold=True)
        the_ws.column_dimensions['C'].width = 9

        # NEXT COLUMN
        col_incr = 3
        the_ws["A1"].offset(row=row_incr, column=col_incr).value = NA_COLUMN_LABEL
        the_ws["A1"].offset(row=row_incr, column=col_incr).alignment = Alignment(horizontal='center')
        the_ws["A1"].offset(row=row_incr, column=col_incr).font = Font(bold=True)
        the_ws.column_dimensions['D'].width = 10

        # NEXT COLUMN
        col_incr = 4
        the_ws["A1"].offset(row=row_incr, column=col_incr).value = MEAN_HEADER
        the_ws["A1"].offset(row=row_incr, column=col_incr).alignment = Alignment(horizontal='center')
        the_ws["A1"].offset(row=row_incr, column=col_incr).font = Font(bold=True)
        the_ws.column_dimensions['E'].width = 12

        # NEXT COLUMN
        col_incr = 5
        the_ws["A1"].offset(row=row_incr, column=col_incr).value = MEDIAN_HEADER
        the_ws["A1"].offset(row=row_incr, column=col_incr).alignment = Alignment(horizontal='center')
        the_ws["A1"].offset(row=row_incr, column=col_incr).font = Font(bold=True)
        the_ws.column_dimensions['F'].width = 12

        # NEXT COLUMN
        col_incr = 6
        the_ws["A1"].offset(row=row_incr, column=col_incr).value = VARIANCE_HEADER
        the_ws["A1"].offset(row=row_incr, column=col_incr).alignment = Alignment(horizontal='center')
        the_ws["A1"].offset(row=row_incr, column=col_incr).font = Font(bold=True)
        the_ws.column_dimensions['G'].width = 15

        # NEXT COLUMN
        col_incr = 7
        the_ws["A1"].offset(row=row_incr, column=col_incr).value = STD_DEV_HEADER
        the_ws["A1"].offset(row=row_incr, column=col_incr).alignment = Alignment(horizontal='center')
        the_ws["A1"].offset(row=row_incr, column=col_incr).font = Font(bold=True)
        the_ws.column_dimensions['H'].width = 12

        # NEXT COLUMN
        col_incr = 8
        the_ws["A1"].offset(row=row_incr, column=col_incr).value = SKEW_HEADER
        the_ws["A1"].offset(row=row_incr, column=col_incr).alignment = Alignment(horizontal='center')
        the_ws["A1"].offset(row=row_incr, column=col_incr).font = Font(bold=True)
        the_ws.column_dimensions['I'].width = 12

        # NEXT COLUMN
        col_incr = 9
        the_ws["A1"].offset(row=row_incr, column=col_incr).value = MIN_HEADER
        the_ws["A1"].offset(row=row_incr, column=col_incr).alignment = Alignment(horizontal='center')
        the_ws["A1"].offset(row=row_incr, column=col_incr).font = Font(bold=True)
        the_ws.column_dimensions['J'].width = 12

        # NEXT COLUMN
        col_incr = 10
        the_ws["A1"].offset(row=row_incr, column=col_incr).value = MAX_HEADER
        the_ws["A1"].offset(row=row_incr, column=col_incr).alignment = Alignment(horizontal='center')
        the_ws["A1"].offset(row=row_incr, column=col_incr).font = Font(bold=True)
        the_ws.column_dimensions['K'].width = 12

        # only generate this header if we're doing a FULL
        if the_type == ANALYZE_DATASET_FULL:
            # NEXT ROW
            col_incr = 11
            the_ws["A1"].offset(row=row_incr, column=col_incr).value = DISTRIBUTION
            the_ws["A1"].offset(row=row_incr, column=col_incr).alignment = Alignment(horizontal='center')
            the_ws["A1"].offset(row=row_incr, column=col_incr).font = Font(bold=True)
            the_ws.column_dimensions['L'].width = 16

        # we need to freeze the top row
        the_ws.freeze_panes = the_ws["B2"]

        # log that we're about to start looping over the floats
        self.logger.debug("About to start adding float items to page.")

        # define the row_count
        row_count = 1

        # loop over the items in the list
        for next_value in the_list:
            # log the first retrieved item
            logging.debug(f"retrieved float [{next_value}]")

            # get the item #
            item_num = count_dict[next_value]

            # write out the item #
            the_ws["A1"].offset(row=row_count, column=0).value = item_num
            the_ws["A1"].offset(row=row_count, column=0).alignment = Alignment(horizontal='center')

            # write out the column name
            the_ws["A1"].offset(row=row_count, column=1).value = next_value
            the_ws["A1"].offset(row=row_count, column=1).alignment = Alignment(horizontal='left')

            # write out the TOTAL count
            the_ws["A1"].offset(row=row_count, column=2).value = self.data_analyzer.the_df[next_value].count()
            the_ws["A1"].offset(row=row_count, column=2).alignment = Alignment(horizontal='center')
            the_ws["A1"].offset(row=row_count, column=2).number_format = FORMAT_NUMBER_COMMA

            # write out the NaN count
            the_ws["A1"].offset(row=row_count, column=3).value = nan_count_dict[next_value]
            the_ws["A1"].offset(row=row_count, column=3).alignment = Alignment(horizontal='center')
            the_ws["A1"].offset(row=row_count, column=3).number_format = FORMAT_NUMBER_COMMA

            # write out mean
            the_ws["A1"].offset(row=row_count, column=4).value = self.data_analyzer.the_df[next_value].mean()
            the_ws["A1"].offset(row=row_count, column=4).alignment = Alignment(horizontal='center')
            the_ws["A1"].offset(row=row_count, column=4).number_format = "#,##0.000"

            # write out median
            the_ws["A1"].offset(row=row_count, column=5).value = self.data_analyzer.the_df[next_value].median()
            the_ws["A1"].offset(row=row_count, column=5).alignment = Alignment(horizontal='center')
            the_ws["A1"].offset(row=row_count, column=5).number_format = "#,##0.000"

            # write out variance
            the_ws["A1"].offset(row=row_count, column=6).value = self.data_analyzer.the_df[next_value].var()
            the_ws["A1"].offset(row=row_count, column=6).alignment = Alignment(horizontal='center')
            the_ws["A1"].offset(row=row_count, column=6).number_format = "#,##0.000"

            # write out variance
            the_ws["A1"].offset(row=row_count, column=7).value = self.data_analyzer.the_df[next_value].std()
            the_ws["A1"].offset(row=row_count, column=7).alignment = Alignment(horizontal='center')
            the_ws["A1"].offset(row=row_count, column=7).number_format = "#,##0.000"

            # write out skew
            the_ws["A1"].offset(row=row_count, column=8).value = self.data_analyzer.the_df[next_value].skew()
            the_ws["A1"].offset(row=row_count, column=8).alignment = Alignment(horizontal='center')
            the_ws["A1"].offset(row=row_count, column=8).number_format = "#,##0.0000"

            # write out min
            the_ws["A1"].offset(row=row_count, column=9).value = self.data_analyzer.the_df[next_value].min()
            the_ws["A1"].offset(row=row_count, column=9).alignment = Alignment(horizontal='center')
            the_ws["A1"].offset(row=row_count, column=9).number_format = FORMAT_NUMBER_COMMA

            # write out max
            the_ws["A1"].offset(row=row_count, column=10).value = self.data_analyzer.the_df[next_value].max()
            the_ws["A1"].offset(row=row_count, column=10).alignment = Alignment(horizontal='center')
            the_ws["A1"].offset(row=row_count, column=10).number_format = FORMAT_NUMBER_COMMA

            if the_type == ANALYZE_DATASET_FULL:
                # write out the distribution
                the_ws["A1"].offset(row=row_count, column=11).value = dist_dict[next_value][DIST_NAME]
                the_ws["A1"].offset(row=row_count, column=11).alignment = Alignment(horizontal='center')

            # increment row_count
            row_count = row_count + 1

    # create float analysis sheet
    def __create_float_field_analysis_sheet__(self, the_index, the_type=None):
        # log that we've been called
        logging.info("A request to generate the float field analysis tab has been made.")

        # what I want to do is loop over the list of floats columns, and grab all the images
        # (in this first run, I'm just doing histograms) for the column and place them horizontally
        # onto the generalized rows associated with the float column.  I will construct a general header
        # for the report, where each float column will have a series of rows that display relevant
        # plots for exploring the data.

        # create the basic analysis tab
        self.the_workbook.create_sheet(index=the_index, title="FLOAT DATA FIELD ANALYSIS")

        # get the sheet references
        the_ws = self.the_workbook.worksheets[the_index]

        # Get the list of FLOAT columns
        the_list = self.data_analyzer.storage[FLOAT64_COLUMN_KEY]

        # get the dict of what row number the field is in.  Required for the header
        count_dict = self.data_analyzer.storage[COLUMN_COUNT_KEY]

        # get the dict of plots
        plot_dict = self.plot_generator.plot_storage

        # retrieve the list of BOOLEAN columns on key BOOL_COLUMN_KEY
        exclusion_list = self.data_analyzer.storage[BOOL_COLUMN_KEY]

        # log the number of floats we have
        logging.info(f"The list of floats is of size [{len(the_list)}]")

        # set the base cell that we start all calculations
        base_cell_addr = "A1"
        base_col_name_addr = "F1:G1"
        base_col_name_value_addr = "H1:L1"
        the_offset = 0
        relative_offset = 0
        previous_offset = 0
        column_offset = 31  # this is the initial value of the offset

        # define the ranges
        cell_range_F1G1 = CellRange(base_col_name_addr)
        cell_range_H1L1 = CellRange(base_col_name_value_addr)

        # define the image reference.
        the_image = None

        # we now need to loop over the columns in the float list
        for the_column in the_list:
            # log the column that we're getting
            logging.debug("column[%s] ", the_column)

            # Write out the top header in row 1
            # COLUMN A - COLUMN # - HEADER
            the_ws[base_cell_addr].offset(row=the_offset, column=0).value = BASIC_ITEM_HEADER
            the_ws[base_cell_addr].offset(row=the_offset, column=0).alignment = Alignment(horizontal='center')
            the_ws[base_cell_addr].offset(row=the_offset, column=0).font = Font(bold=True)
            the_ws[base_cell_addr].offset(row=the_offset, column=0).border = \
                Border(top=BORDER_THICK, left=BORDER_THICK, right=BORDER_THICK, bottom=BORDER_THICK)
            the_ws.column_dimensions['A'].width = 24

            # COLUMN B - COLUMN # - Actual Value
            the_ws[base_cell_addr].offset(row=the_offset, column=1).value = count_dict[the_column]
            the_ws[base_cell_addr].offset(row=the_offset, column=1).alignment = Alignment(horizontal='center')
            the_ws[base_cell_addr].offset(row=the_offset, column=1).font = Font(bold=True)
            the_ws[base_cell_addr].offset(row=the_offset, column=1).border = \
                Border(top=BORDER_THICK, left=BORDER_THICK, right=BORDER_THICK, bottom=BORDER_THICK)
            the_ws.column_dimensions['B'].width = 9

            # COLUMN F1:G1 - COLUMN NAME - HEADER
            the_ws.merge_cells(start_row=the_offset + 1, start_column=6, end_row=the_offset + 1, end_column=7)
            the_ws[base_cell_addr].offset(row=the_offset, column=5).value = BASIC_COLUMN_HEADER
            the_ws[base_cell_addr].offset(row=the_offset, column=5).alignment = Alignment(horizontal='center')
            the_ws[base_cell_addr].offset(row=the_offset, column=5).font = Font(bold=True)

            # offset the range
            cell_range_F1G1.shift(col_shift=0, row_shift=relative_offset)

            # set the border
            self.set_border(the_ws, cell_range_F1G1)

            # COLUMN H1:L1 - COLUMN NAME - Actual Value
            the_ws.merge_cells(start_row=the_offset + 1, start_column=8, end_row=the_offset + 1, end_column=12)
            the_ws[base_cell_addr].offset(row=the_offset, column=7).value = the_column
            the_ws[base_cell_addr].offset(row=the_offset, column=7).alignment = Alignment(horizontal='left')
            the_ws[base_cell_addr].offset(row=the_offset, column=7).font = Font(bold=True)

            # offset the range
            cell_range_H1L1.shift(col_shift=0, row_shift=relative_offset)

            # set the border
            self.set_border(the_ws, cell_range_H1L1)

            # ***************************************************************************
            #                                HISTOGRAMS
            # ***************************************************************************

            # log the image we're going to import
            logging.debug(f"About to insert image [{plot_dict[the_column][PLOT_TYPE_HIST]}]")

            # create an instance of the image
            the_image = Image(plot_dict[the_column][PLOT_TYPE_HIST])
            the_image.anchor = the_ws[base_cell_addr].offset(row=the_offset + 2, column=1).coordinate

            # log where we're putting the image
            logging.debug("Image anchor is [%s]",
                          the_ws[base_cell_addr].offset(row=the_offset + 2, column=1).coordinate)

            # add the image to the worksheet
            the_ws.add_image(the_image)

            # ***************************************************************************
            #                                 BOX PLOTS
            # ***************************************************************************

            # log the image we're going to import
            logging.debug(f"About to insert image [{plot_dict[the_column][PLOT_TYPE_BOX_PLOT]}]")

            # create an instance of the image
            the_image = Image(plot_dict[the_column][PLOT_TYPE_BOX_PLOT])
            the_image.anchor = the_ws[base_cell_addr].offset(row=the_offset + 2, column=11).coordinate

            # log where we're putting the image
            logging.debug(f"Image anchor is ["
                          f"{the_ws[base_cell_addr].offset(row=the_offset + 2, column=11).coordinate}]")

            # add the image to the worksheet
            the_ws.add_image(the_image)

            # log all the offset values
            logging.debug(f"prior to adjustment, the_offset[{the_offset}] "
                          f"relative_offset[{relative_offset}] previous_offset[{previous_offset}]")

            # ***************************************************************************
            #                                 QQ PLOTS
            # ***************************************************************************

            # log the image we're going to import
            logging.debug(f"About to insert image [{plot_dict[the_column][PLOT_TYPE_Q_Q_PLOT]}]")

            # create an instance of the image
            the_image = Image(plot_dict[the_column][PLOT_TYPE_Q_Q_PLOT])
            the_image.anchor = the_ws[base_cell_addr].offset(row=the_offset + 2, column=21).coordinate

            # log where we're putting the image
            logging.debug(f"Image anchor is ["
                          f"{the_ws[base_cell_addr].offset(row=the_offset + 2, column=21).coordinate}]")

            # add the image to the worksheet
            the_ws.add_image(the_image)

            # log all the offset values
            logging.debug(f"prior to adjustment, the_offset[{the_offset}] "
                          f"relative_offset[{relative_offset}] previous_offset[{previous_offset}]")

            # ***************************************************************************
            #                                 JOINT PLOTS
            # ***************************************************************************
            if the_type == ANALYZE_DATASET_FULL:
                # get the stat generator
                the_sg = self.stat_generator

                # get the list of tuples for our current column, excluding boolean fields
                tuple_list = the_sg.filter_tuples_by_column(the_sg.all_corr_storage, the_column, exclusion_list)

                # log the tuple_list returned
                self.logger.debug(f"the tuple_list for [{the_column}] is {tuple_list}")

                # loop over all the tuples
                for the_tuple in tuple_list:
                    # log the image we're going to import
                    logging.debug(f"for [{the_column}] inserting image ["
                                  f"{plot_dict[PLOT_TYPE_JOINT_PLOT][the_tuple]}]")

                    # create an instance of the image
                    the_image = Image(plot_dict[PLOT_TYPE_JOINT_PLOT][the_tuple])
                    the_image.anchor = the_ws[base_cell_addr].offset(row=the_offset + 2,
                                                                     column=column_offset).coordinate

                    # log where we're putting the image
                    logging.debug(f"Image anchor is [{the_image.anchor}]")

                    # add the image to the worksheet
                    the_ws.add_image(the_image)

                    # increment the column offset
                    column_offset = column_offset + 10

            # ***************************************************************************
            #                                 END OF PLOTS
            # ***************************************************************************

            # force the offset to MINIMUM_OFFSET
            the_offset = the_offset + MINIMUM_OFFSET + 5

            # capture the relative offset for this loop
            relative_offset = the_offset - previous_offset

            # reset the previous_offset
            previous_offset = the_offset

            # reset the column offset
            column_offset = 31

            # log all the offset values
            logging.debug(f"At end of function, the_offset[{the_offset}] "
                          f"relative_offset[{relative_offset}] previous_offset[{previous_offset}]")

        # log that we are complete.
        logging.debug("Complete with generation of graphs for OBJECT data analysis.")

    # create boolean analysis sheet
    def __create_boolean_analysis_sheet__(self, the_index):
        # log that we've been called
        self.logger.info("A request to generate the boolean data tab has been made.")

        # create the basic analysis tab
        self.the_workbook.create_sheet(index=the_index, title="BOOLEAN DATA ANALYSIS")

        # get the sheet
        the_ws = self.the_workbook.worksheets[the_index]

        # get the list of booleans
        the_list = self.data_analyzer.storage[BOOL_COLUMN_KEY]

        # log how many we have
        self.logger.info(f"There list of booleans is of size [{len(the_list)}]")

        # get the dict of unique values key is column, value is list of unique values
        unique_values = self.data_analyzer.storage[UNIQUE_COLUMN_VALUES]

        # get the dict of what row number the field is in
        count_dict = self.data_analyzer.storage[COLUMN_COUNT_KEY]

        # get dict of true/false counts stored at BOOL_COLUMN_COUNT_KEY
        bool_col_count_dict = self.data_analyzer.storage[BOOL_COLUMN_COUNT_KEY]

        # log that we're about to create the headers
        self.logger.debug("About to start creating the header row.")

        # let's select cell A1 to start writing out the headers
        the_ws["A1"].value = BASIC_ITEM_HEADER
        the_ws["A1"].alignment = Alignment(horizontal='center')
        the_ws["A1"].font = Font(bold=True)
        the_ws.column_dimensions['A'].width = 9

        the_ws["B1"].value = BASIC_COLUMN_HEADER
        the_ws["B1"].font = Font(bold=True)
        the_ws.column_dimensions['B'].width = 20

        the_ws["C1"].value = BASIC_COUNT_HEADER
        the_ws["C1"].alignment = Alignment(horizontal='center')
        the_ws["C1"].font = Font(bold=True)
        the_ws.column_dimensions['C'].width = 9

        the_ws["D1"].value = VALUES_ACTUAL
        the_ws["D1"].alignment = Alignment(horizontal='center')
        the_ws["D1"].font = Font(bold=True)
        the_ws.column_dimensions['D'].width = 20

        the_ws["E1"].value = TRUE_COUNT
        the_ws["E1"].alignment = Alignment(horizontal='center')
        the_ws["E1"].font = Font(bold=True)
        the_ws.column_dimensions['E'].width = 15

        the_ws["F1"].value = FALSE_COUNT
        the_ws["F1"].alignment = Alignment(horizontal='center')
        the_ws["F1"].font = Font(bold=True)
        the_ws.column_dimensions['F'].width = 15

        # we need to freeze the top row
        the_ws.freeze_panes = the_ws["B2"]

        # log that we're about to start looping over the booleans
        self.logger.debug("About to start adding boolean items to page.")

        # define the row_count
        row_count = 1

        # loop over the items in the list
        for next_value in the_list:
            # log the first retrieved item
            logging.debug(f"retrieved boolean [{next_value}]")

            # write out the item #
            the_ws["A1"].offset(row=row_count, column=0).value = count_dict[next_value]
            the_ws["A1"].offset(row=row_count, column=0).alignment = Alignment(horizontal='center')

            # write out the column name
            the_ws["A1"].offset(row=row_count, column=1).value = next_value
            the_ws["A1"].offset(row=row_count, column=1).alignment = Alignment(horizontal='left')

            # write out the count
            the_ws["A1"].offset(row=row_count, column=2).value = self.data_analyzer.the_df[next_value].count()
            the_ws["A1"].offset(row=row_count, column=2).alignment = Alignment(horizontal='center')
            the_ws["A1"].offset(row=row_count, column=2).number_format = FORMAT_NUMBER_COMMA

            # write out unique values
            the_ws["A1"].offset(row=row_count, column=3).value = str(unique_values[next_value])
            the_ws["A1"].offset(row=row_count, column=3).alignment = Alignment(horizontal='center')

            # write out TRUE count
            the_ws["A1"].offset(row=row_count, column=4).value = bool_col_count_dict[next_value]['True']
            the_ws["A1"].offset(row=row_count, column=4).alignment = Alignment(horizontal='center')

            # write out FALSE count
            the_ws["A1"].offset(row=row_count, column=5).value = bool_col_count_dict[next_value]['False']
            the_ws["A1"].offset(row=row_count, column=5).alignment = Alignment(horizontal='center')

            # increment row_count
            row_count = row_count + 1

    # create boolean count sheet
    def __create_boolean_count_sheet__(self, the_index, the_type=None):
        # log that we've been called
        logging.info(f"A request to generate the boolean data analysis tab for {the_type}.")

        # create the basic analysis tab
        self.the_workbook.create_sheet(index=the_index, title="BOOLEAN DATA FIELD ANALYSIS")

        # get the sheet
        the_ws = self.the_workbook.worksheets[the_index]

        # Get the list of BOOLEAN column strings.  This list has been stripped of booleans
        the_list = self.data_analyzer.storage[BOOL_COLUMN_KEY]

        # log the number of objects we have
        logging.debug(f"The list of booleans is of size [{len(the_list)}]")

        # get the dict of total count of values in each column.  Key is column, value is total count
        total_count_dict = self.data_analyzer.storage[COLUMN_TOTAL_COUNT_KEY]

        # get the dict of what row number the field is in
        count_dict = self.data_analyzer.storage[COLUMN_COUNT_KEY]

        # get the dict of plots
        plot_dict = self.plot_generator.plot_storage

        # create a list of integer, float, and object columns.  This will be used for box plots
        the_column_list = self.data_analyzer.storage[INT64_COLUMN_KEY] + self.data_analyzer.storage[FLOAT64_COLUMN_KEY]

        # set the base cell that we start all calculations
        base_cell_addr = "A1"
        base_col_name_addr = "F1:G1"
        base_col_name_value_addr = "H1:L1"
        the_offset = 0
        previous_offset = 0
        relative_offset = 0
        column_previous_offset = 14

        # define the ranges
        cell_range_F1G1 = CellRange(base_col_name_addr)
        cell_range_H1L1 = CellRange(base_col_name_value_addr)

        # we now need to loop over the columns in the object list
        for the_column in the_list:
            # # get the ratio
            # the_col_ratio = ratio_dict[the_column]

            # log the column that we're getting
            logging.info(f"current column[{the_column}s].")

            # log that we're generating graphs
            logging.debug(f"Generating graphs for column[{the_column}] with offset[{the_offset}s]")

            # get the list of valid values
            unique_col_value_list = list(self.data_analyzer.storage[BOOL_COLUMN_COUNT_KEY][the_column].keys())

            # get the total number of values in this column
            unique_col_count = total_count_dict[the_column]

            # cumulative total field count
            cum_field_count = 0

            # Write out the top header in row 1
            # COLUMN A - COLUMN # - HEADER
            the_ws[base_cell_addr].offset(row=the_offset, column=0).value = BASIC_ITEM_HEADER
            the_ws[base_cell_addr].offset(row=the_offset, column=0).alignment = Alignment(horizontal='center')
            the_ws[base_cell_addr].offset(row=the_offset, column=0).font = Font(bold=True)
            the_ws[base_cell_addr].offset(row=the_offset, column=0).border = \
                Border(top=BORDER_THICK, left=BORDER_THICK, right=BORDER_THICK, bottom=BORDER_THICK)

            # COLUMN B - COLUMN # - Actual Value
            the_ws[base_cell_addr].offset(row=the_offset, column=1).value = count_dict[the_column]
            the_ws[base_cell_addr].offset(row=the_offset, column=1).alignment = Alignment(horizontal='center')
            the_ws[base_cell_addr].offset(row=the_offset, column=1).font = Font(bold=True)
            the_ws[base_cell_addr].offset(row=the_offset, column=1).border = \
                Border(top=BORDER_THICK, left=BORDER_THICK, right=BORDER_THICK, bottom=BORDER_THICK)

            # COLUMN F1:G1 - COLUMN NAME - HEADER
            the_ws.merge_cells(start_row=the_offset + 1, start_column=6, end_row=the_offset + 1, end_column=7)
            the_ws[base_cell_addr].offset(row=the_offset, column=5).value = BASIC_COLUMN_HEADER
            the_ws[base_cell_addr].offset(row=the_offset, column=5).alignment = Alignment(horizontal='center')
            the_ws[base_cell_addr].offset(row=the_offset, column=5).font = Font(bold=True)

            # offset the range
            cell_range_F1G1.shift(col_shift=0, row_shift=relative_offset)

            # set the border
            self.set_border(the_ws, cell_range_F1G1)

            # COLUMN H1:L1 - COLUMN NAME - Actual Value
            the_ws.merge_cells(start_row=the_offset + 1, start_column=8, end_row=the_offset + 1, end_column=12)
            the_ws[base_cell_addr].offset(row=the_offset, column=7).value = the_column
            the_ws[base_cell_addr].offset(row=the_offset, column=7).alignment = Alignment(horizontal='left')
            the_ws[base_cell_addr].offset(row=the_offset, column=7).font = Font(bold=True)

            # offset the range
            cell_range_H1L1.shift(col_shift=0, row_shift=relative_offset)

            # set the border
            self.set_border(the_ws, cell_range_H1L1)

            # Write out the column headers in row 3
            # COLUMN A - COLUMN VALUE Header
            the_ws[base_cell_addr].offset(row=the_offset + 2, column=0).value = OBJ_DATA_FIELD_COLUMN_VAL
            the_ws[base_cell_addr].offset(row=the_offset + 2, column=0).alignment = Alignment(horizontal='left')
            the_ws[base_cell_addr].offset(row=the_offset + 2, column=0).font = Font(bold=True)
            the_ws.column_dimensions['A'].width = 24

            # COLUMN B - COUNT Header
            the_ws[base_cell_addr].offset(row=the_offset + 2, column=1).value = OBJ_DATA_FIELD_COUNT
            the_ws[base_cell_addr].offset(row=the_offset + 2, column=1).alignment = Alignment(horizontal='center')
            the_ws[base_cell_addr].offset(row=the_offset + 2, column=1).font = Font(bold=True)
            the_ws.column_dimensions['B'].width = 9

            # COLUMN C - COLUMN % Header
            the_ws[base_cell_addr].offset(row=the_offset + 2, column=2).value = OBJ_DATA_FIELD_COLUMN_PCT
            the_ws[base_cell_addr].offset(row=the_offset + 2, column=2).alignment = Alignment(horizontal='center')
            the_ws[base_cell_addr].offset(row=the_offset + 2, column=2).font = Font(bold=True)
            the_ws.column_dimensions['C'].width = 12

            # loop over the list of unique values for column in "the_column" variable
            for unique_value in unique_col_value_list:
                # log the unique column value
                logging.debug(f"unique column value[{unique_value}] for column[{the_column}]")

                unique_value_count = self.data_analyzer.storage[BOOL_COLUMN_COUNT_KEY][the_column][unique_value]

                # log what we're seeing.
                logging.debug(f"The count is [{unique_value_count}] for unique value [{unique_value}]")

                # calculate the current cumulative field count
                cum_field_count = cum_field_count + unique_value_count

                # COLUMN A - COLUMN VALUE Actual Value
                the_ws[base_cell_addr].offset(row=the_offset + 3, column=0).value = unique_value
                the_ws[base_cell_addr].offset(row=the_offset + 3, column=0).alignment = Alignment(horizontal='left')

                # COLUMN B - COUNT Actual Value
                the_ws[base_cell_addr].offset(row=the_offset + 3, column=1).value = unique_value_count
                the_ws[base_cell_addr].offset(row=the_offset + 3, column=1).alignment = Alignment(
                    horizontal='center')
                the_ws[base_cell_addr].offset(row=the_offset + 3, column=1).number_format = FORMAT_NUMBER_COMMA

                # COLUMN C - COLUMN % Actual Value
                the_ws[base_cell_addr].offset(row=the_offset + 3, column=2).value = \
                    unique_value_count / unique_col_count
                the_ws[base_cell_addr].offset(row=the_offset + 3, column=2).alignment = Alignment(
                    horizontal='center')
                the_ws[base_cell_addr].offset(row=the_offset + 3, column=2).number_format = FORMAT_PERCENTAGE

                # increment the offset within a specific section (or values for a specific column)
                the_offset = the_offset + 1

                # log the value of the offset
                logging.debug(f"For value [{unique_value}] the offset is [{the_offset}]")

            # increment the offset for a section jump (each column of unique values is a section)
            the_offset = the_offset + DEFAULT_ROW_HEADER_SPACER + DEFAULT_ROW_SPACER

            # make sure the total offset is greater than MINIMUM_OFFSET
            if the_offset - previous_offset < MINIMUM_OFFSET:
                # log the offset values
                logging.debug(f"The initial offset is [{the_offset}] previous offset is [{previous_offset}]")

                # capture the relative offset for this loop
                relative_offset = MINIMUM_OFFSET

                # force the offset to MINIMUM_OFFSET
                the_offset = the_offset + MINIMUM_OFFSET - (the_offset - previous_offset)

                # log the updated offset
                logging.debug(f"The updated offset is now [{the_offset}] "
                              f"the width is [{the_offset - previous_offset}]")
            else:
                # capture the relative offset for this loop
                relative_offset = the_offset - previous_offset

            # log the relative offset
            logging.debug(f"The relative offset is [{relative_offset}]")

            # ***************************************************************************
            #                                 BAR CHART PLOTS
            # ***************************************************************************

            # log the image we're going to import
            logging.debug(f"About to insert bar chart image [{plot_dict[the_column][PLOT_TYPE_BAR_CHART]}]")

            # create an instance of the image
            the_image = Image(plot_dict[the_column][PLOT_TYPE_BAR_CHART])
            the_image.anchor = the_ws[base_cell_addr].offset(row=previous_offset + 2, column=4).coordinate

            # log where we're putting the image
            logging.debug(f"Image anchor is [{the_image.anchor}]")

            # add the image to the worksheet
            the_ws.add_image(the_image)

            # log all the offset values
            logging.debug(f"prior to adjustment, the_offset[{the_offset}] "
                          f"relative_offset[{relative_offset}] previous_offset[{previous_offset}]")

            if the_type == ANALYZE_DATASET_FULL:
                # ***************************************************************************
                #                                 COUNT PLOTS
                # ***************************************************************************

                # create an exclusion list of FLOAT64_COLUMN_KEY and INT64_COLUMN_KEY
                exclusion_list = self.data_analyzer.storage[FLOAT64_COLUMN_KEY] + \
                                 self.data_analyzer.storage[INT64_COLUMN_KEY]

                # invoke the method
                tuple_list = self.stat_generator.filter_tuples_by_column(
                    the_storage=self.stat_generator.chi_square_results,
                    the_column=the_column,
                    exclusion_list=exclusion_list)

                # log the tuple_list returned
                self.logger.info(f"for the boolean count plots, the tuple_list is {tuple_list}")

                # loop over all the tuples
                for the_tuple in tuple_list:
                    # log the image we're going to import
                    logging.info(f"for [{the_column}][{the_tuple}] inserting image ["
                                 f"{plot_dict[PLOT_TYPE_BIVARIATE_COUNT][the_tuple]}]")

                    # create an instance of the image
                    the_image = Image(plot_dict[PLOT_TYPE_BIVARIATE_COUNT][the_tuple])
                    the_image.anchor = the_ws[base_cell_addr].offset(row=previous_offset + 2,
                                                                     column=column_previous_offset).coordinate

                    # log where we're putting the image
                    logging.info(f"\tImage anchor is [{the_image.anchor}]")

                    # add the image to the worksheet
                    the_ws.add_image(the_image)

                    # increment the column offset
                    column_previous_offset = column_previous_offset + 11

                # ***************************************************************************
                #                                 BOX PLOTS
                # ***************************************************************************
                # start to loop over the_column_list
                for second_column in the_column_list:
                    # create the tuple
                    the_box_plot_tuple = (the_column, second_column)

                    # log the tuple_list returned
                    self.logger.info(f"for the boolean box plots, the current plot is {the_box_plot_tuple}.")

                    # create an instance of the image
                    the_image = Image(plot_dict[PLOT_TYPE_BIVARIATE_BOX][the_box_plot_tuple])
                    the_image.anchor = the_ws[base_cell_addr].offset(row=previous_offset + 2,
                                                                     column=column_previous_offset).coordinate

                    # log where we're putting the image
                    logging.info(f"\tImage anchor is [{the_image.anchor}]")

                    # add the image to the worksheet
                    the_ws.add_image(the_image)

                    # increment the column offset
                    column_previous_offset = column_previous_offset + 11

                # we need to reset the column offset back to initial state.
                column_previous_offset = 14

            # reset the previous_offset - moves us in y direction.
            previous_offset = the_offset

        # log that we are complete.
        logging.debug("Complete with generation of graphs for OBJECT data analysis.")

    # create object count sheet
    def __create_object_count_sheet__(self, the_index, the_type=None):
        # log that we've been called
        logging.info("A request to generate the object data analysis tab has been made.")

        # create the basic analysis tab
        self.the_workbook.create_sheet(index=the_index, title="OBJECT DATA FIELD ANALYSIS")

        # get the sheet
        the_ws = self.the_workbook.worksheets[the_index]

        # Get the list of OBJECT column strings.  This list has been stripped of booleans
        the_list = self.data_analyzer.storage[OBJECT_COLUMN_KEY]

        # log the number of objects we have
        logging.info(f"The list of objects is of size [{len(the_list)}]")

        # get the dictionary of unique values.  Key - column name, value = list of unique values
        unique_values = self.data_analyzer.storage[UNIQUE_COLUMN_VALUES]

        # get the dict of what row number the field is in
        count_dict = self.data_analyzer.storage[COLUMN_COUNT_KEY]

        # get the dict of total count of values in each column.  Key is column, value is total count
        total_count_dict = self.data_analyzer.storage[COLUMN_TOTAL_COUNT_KEY]

        # get the dict of ratios. key is column, value is ratio
        ratio_dict = self.data_analyzer.storage[UNIQUE_COLUMN_RATIOS]

        # get the dict of counts for values for object column.  The key is column, the value is a dict.
        obj_column_count_dict = self.data_analyzer.storage[OBJECT_COLUMN_COUNT_KEY]

        # get the dict of plots
        plot_dict = self.plot_generator.plot_storage

        # create a list of integer, float, and boolean columns.  This will be used for box plots.
        the_column_list = self.data_analyzer.storage[INT64_COLUMN_KEY] + self.data_analyzer.storage[FLOAT64_COLUMN_KEY]

        # set the base cell that we start all calculations
        base_cell_addr = "A1"
        base_col_name_addr = "F1:G1"
        base_col_name_value_addr = "H1:L1"
        the_offset = 0
        previous_offset = 0
        relative_offset = 0
        column_previous_offset = 16

        # define the ranges
        cell_range_F1G1 = CellRange(base_col_name_addr)
        cell_range_H1L1 = CellRange(base_col_name_value_addr)

        # we now need to loop over the columns in the object list
        for the_column in the_list:
            # get the ratio
            the_col_ratio = ratio_dict[the_column]

            # log the column that we're getting
            logging.debug(f"column[{the_column}s] ratio[{the_col_ratio}]")

            # We need to check if the unique rate on key UNIQUE_COLUMN_RATIOS is 1.0, meaning all values are unique.
            if the_col_ratio >= 0.05:
                # log that we're not creating a section for this column
                logging.debug(f"Skipping graphs for column [{the_column}]")
            # we can display results
            else:
                # log that we're generating graphs
                logging.debug(f"Generating graphs for column[{the_column}] with offset[{the_offset}s]")

                # get the unique values list for this column
                unique_col_value_list = unique_values[the_column]

                # get the total number of values in this column
                unique_col_count = total_count_dict[the_column]

                # cumulative total field count
                cum_field_count = 0

                # total field count
                total_field_count = total_count_dict[the_column]

                # Write out the top header in row 1
                # COLUMN A - COLUMN # - HEADER
                the_ws[base_cell_addr].offset(row=the_offset, column=0).value = BASIC_ITEM_HEADER
                the_ws[base_cell_addr].offset(row=the_offset, column=0).alignment = Alignment(horizontal='center')
                the_ws[base_cell_addr].offset(row=the_offset, column=0).font = Font(bold=True)
                the_ws[base_cell_addr].offset(row=the_offset, column=0).border = \
                    Border(top=BORDER_THICK, left=BORDER_THICK, right=BORDER_THICK, bottom=BORDER_THICK)

                # COLUMN B - COLUMN # - Actual Value
                the_ws[base_cell_addr].offset(row=the_offset, column=1).value = count_dict[the_column]
                the_ws[base_cell_addr].offset(row=the_offset, column=1).alignment = Alignment(horizontal='center')
                the_ws[base_cell_addr].offset(row=the_offset, column=1).font = Font(bold=True)
                the_ws[base_cell_addr].offset(row=the_offset, column=1).border = \
                    Border(top=BORDER_THICK, left=BORDER_THICK, right=BORDER_THICK, bottom=BORDER_THICK)

                # COLUMN F1:G1 - COLUMN NAME - HEADER
                the_ws.merge_cells(start_row=the_offset + 1, start_column=6, end_row=the_offset + 1, end_column=7)
                the_ws[base_cell_addr].offset(row=the_offset, column=5).value = BASIC_COLUMN_HEADER
                the_ws[base_cell_addr].offset(row=the_offset, column=5).alignment = Alignment(horizontal='center')
                the_ws[base_cell_addr].offset(row=the_offset, column=5).font = Font(bold=True)

                # offset the range
                cell_range_F1G1.shift(col_shift=0, row_shift=relative_offset)

                # set the border
                self.set_border(the_ws, cell_range_F1G1)

                # COLUMN H1:L1 - COLUMN NAME - Actual Value
                the_ws.merge_cells(start_row=the_offset + 1, start_column=8, end_row=the_offset + 1, end_column=12)
                the_ws[base_cell_addr].offset(row=the_offset, column=7).value = the_column
                the_ws[base_cell_addr].offset(row=the_offset, column=7).alignment = Alignment(horizontal='left')
                the_ws[base_cell_addr].offset(row=the_offset, column=7).font = Font(bold=True)

                # offset the range
                cell_range_H1L1.shift(col_shift=0, row_shift=relative_offset)

                # set the border
                self.set_border(the_ws, cell_range_H1L1)

                # Write out the column headers in row 3
                # COLUMN A - COLUMN VALUE Header
                the_ws[base_cell_addr].offset(row=the_offset + 2, column=0).value = OBJ_DATA_FIELD_COLUMN_VAL
                the_ws[base_cell_addr].offset(row=the_offset + 2, column=0).alignment = Alignment(horizontal='left')
                the_ws[base_cell_addr].offset(row=the_offset + 2, column=0).font = Font(bold=True)
                the_ws.column_dimensions['A'].width = 24

                # COLUMN B - COUNT Header
                the_ws[base_cell_addr].offset(row=the_offset + 2, column=1).value = OBJ_DATA_FIELD_COUNT
                the_ws[base_cell_addr].offset(row=the_offset + 2, column=1).alignment = Alignment(horizontal='center')
                the_ws[base_cell_addr].offset(row=the_offset + 2, column=1).font = Font(bold=True)
                the_ws.column_dimensions['B'].width = 9

                # COLUMN C - COLUMN % Header
                the_ws[base_cell_addr].offset(row=the_offset + 2, column=2).value = OBJ_DATA_FIELD_COLUMN_PCT
                the_ws[base_cell_addr].offset(row=the_offset + 2, column=2).alignment = Alignment(horizontal='center')
                the_ws[base_cell_addr].offset(row=the_offset + 2, column=2).font = Font(bold=True)
                the_ws.column_dimensions['C'].width = 12

                # COLUMN D - CUM % Header
                the_ws[base_cell_addr].offset(row=the_offset + 2, column=3).value = OBJ_DATA_FIELD_CUM_PCT
                the_ws[base_cell_addr].offset(row=the_offset + 2, column=3).alignment = Alignment(horizontal='center')
                the_ws[base_cell_addr].offset(row=the_offset + 2, column=3).font = Font(bold=True)
                the_ws.column_dimensions['D'].width = 12

                # loop over the list of unique values for column in "the_column" variable
                for unique_value in unique_col_value_list:
                    # log the unique column value
                    logging.debug(f"unique column value[{unique_value}] for column[{the_column}]")

                    # get the count for current field
                    unique_value_count = obj_column_count_dict[the_column][unique_value]

                    # log what we're seeing.
                    logging.debug(f"The count is [{unique_value_count}] for unique value [{unique_value}]")

                    # calculate the current cumulative field count
                    cum_field_count = cum_field_count + unique_value_count

                    # COLUMN A - COLUMN VALUE Actual Value
                    the_ws[base_cell_addr].offset(row=the_offset + 3, column=0).value = unique_value
                    the_ws[base_cell_addr].offset(row=the_offset + 3, column=0).alignment = Alignment(horizontal='left')

                    # COLUMN B - COUNT Actual Value
                    the_ws[base_cell_addr].offset(row=the_offset + 3, column=1).value = unique_value_count
                    the_ws[base_cell_addr].offset(row=the_offset + 3, column=1).alignment = Alignment(
                        horizontal='center')
                    the_ws[base_cell_addr].offset(row=the_offset + 3, column=1).number_format = FORMAT_NUMBER_COMMA

                    # COLUMN C - COLUMN % Actual Value
                    the_ws[base_cell_addr].offset(row=the_offset + 3, column=2).value = \
                        unique_value_count / unique_col_count
                    the_ws[base_cell_addr].offset(row=the_offset + 3, column=2).alignment = Alignment(
                        horizontal='center')
                    the_ws[base_cell_addr].offset(row=the_offset + 3, column=2).number_format = FORMAT_PERCENTAGE

                    # COLUMN D - CUM % Actual Value
                    the_ws[base_cell_addr].offset(row=the_offset + 3,
                                                  column=3).value = cum_field_count / total_field_count
                    the_ws[base_cell_addr].offset(row=the_offset + 3, column=3).alignment = Alignment(
                        horizontal='center')
                    the_ws[base_cell_addr].offset(row=the_offset + 3, column=3).number_format = FORMAT_PERCENTAGE

                    # increment the offset within a specific section (or values for a specific column)
                    the_offset = the_offset + 1

                    # log the value of the offset
                    logging.debug(f"For value [{unique_value}] the offset is [{the_offset}]")

                # increment the offset for a section jump (each column of unique values is a section)
                the_offset = the_offset + DEFAULT_ROW_HEADER_SPACER + DEFAULT_ROW_SPACER

                # make sure the total offset is greater than MINIMUM_OFFSET
                if the_offset - previous_offset < MINIMUM_OFFSET:
                    # log the offset values
                    logging.debug(f"The initial offset is [{the_offset}] previous offset is [{previous_offset}]")

                    # capture the relative offset for this loop
                    relative_offset = MINIMUM_OFFSET

                    # force the offset to MINIMUM_OFFSET
                    the_offset = the_offset + MINIMUM_OFFSET - (the_offset - previous_offset)

                    # log the updated offset
                    logging.debug(f"The updated offset is now [{the_offset}] "
                                  f"the width is [{the_offset - previous_offset}]")
                else:
                    # capture the relative offset for this loop
                    relative_offset = the_offset - previous_offset

                # log the relative offset
                logging.debug(f"The relative offset is [{relative_offset}]")

                # ***************************************************************************
                #                                 BAR CHART PLOTS
                # ***************************************************************************

                # log the image we're going to import
                logging.debug(f"About to insert image [{plot_dict[the_column][PLOT_TYPE_BAR_CHART]}]")

                # create an instance of the image
                the_image = Image(plot_dict[the_column][PLOT_TYPE_BAR_CHART])
                the_image.anchor = the_ws[base_cell_addr].offset(row=previous_offset + 2, column=5).coordinate

                # log where we're putting the image
                logging.debug(f"Image anchor is [{the_image.anchor}]")

                # add the image to the worksheet
                the_ws.add_image(the_image)

                # log all the offset values
                logging.debug(f"prior to adjustment, the_offset[{the_offset}] "
                              f"relative_offset[{relative_offset}] previous_offset[{previous_offset}]")

                # make sure we are in a full analysis
                if the_type == ANALYZE_DATASET_FULL:
                    # ***************************************************************************
                    #                                 COUNT PLOTS
                    # ***************************************************************************

                    # get the list of OBJECT tuples.  This is the entire list, and not specific to the
                    # current value of the_column
                    tuple_list = self.stat_generator.get_list_of_variable_relationships_of_type(OBJECT_COLUMN_KEY)

                    # we need to additionally filter the tuple list by just the_column
                    tuple_list = get_tuples_from_list_with_specific_field(tuple_list, the_column)

                    # log the tuple_list returned
                    self.logger.debug(f"the tuple_list is {tuple_list}")

                    # loop over all the tuples
                    for the_tuple in tuple_list:
                        # log the image we're going to import
                        logging.debug(f"for [{the_column}] inserting image ["
                                      f"{plot_dict[PLOT_TYPE_BIVARIATE_COUNT][the_tuple]}]")

                        # create an instance of the image
                        the_image = Image(plot_dict[PLOT_TYPE_BIVARIATE_COUNT][the_tuple])
                        the_image.anchor = the_ws[base_cell_addr].offset(row=previous_offset + 2,
                                                                         column=column_previous_offset).coordinate

                        # log where we're putting the image
                        logging.debug(f"Image anchor is [{the_image.anchor}]")

                        # add the image to the worksheet
                        the_ws.add_image(the_image)

                        # increment the column offset
                        column_previous_offset = column_previous_offset + 11

                    # ***************************************************************************
                    #                                 BOX PLOTS
                    # ***************************************************************************

                    # start to loop over the_column_list
                    for second_column in the_column_list:
                        # create the tuple
                        the_box_plot_tuple = (the_column, second_column)

                        # log the tuple_list returned
                        self.logger.info(f"for the OBJECT box plots, the current plot is {the_box_plot_tuple}.")

                        # create an instance of the image
                        the_image = Image(plot_dict[PLOT_TYPE_BIVARIATE_BOX][the_box_plot_tuple])
                        the_image.anchor = the_ws[base_cell_addr].offset(row=previous_offset + 2,
                                                                         column=column_previous_offset).coordinate

                        # log where we're putting the image
                        logging.info(f"\tImage anchor is [{the_image.anchor}]")

                        # add the image to the worksheet
                        the_ws.add_image(the_image)

                        # increment the column offset
                        column_previous_offset = column_previous_offset + 11

                    # we need to reset the column offset back to initial state.
                    column_previous_offset = 16

                # reset the previous_offset - moves us in y direction.
                previous_offset = the_offset

        # log that we are complete.
        logging.debug("Complete with generation of graphs for OBJECT data analysis.")

    # create object analysis sheet
    def __create_object_analysis_sheet__(self, the_index):
        # log that we've been called
        self.logger.info("A request to generate the object data tab has been made.")

        # create the basic analysis tab
        self.the_workbook.create_sheet(index=the_index, title="OBJECT DATA ANALYSIS")

        # get the sheet
        the_ws = self.the_workbook.worksheets[the_index]

        # get the list of objects
        the_list = self.data_analyzer.storage[OBJECT_COLUMN_KEY]

        # log the number of objects we have
        self.logger.info(f"The list of objects is of size [{len(the_list)}]")

        # get the list of unique values for each column
        unique_values = self.data_analyzer.storage[UNIQUE_COLUMN_VALUES]

        # get the dict of what row number the field is in
        count_dict = self.data_analyzer.storage[COLUMN_COUNT_KEY]

        # get the flag whether or not we've determined that each value is unique
        unique_column = self.data_analyzer.storage[UNIQUE_COLUMN_FLAG]

        # get the dict of NaN count per column
        nan_count_dict = self.data_analyzer.storage[COLUMN_NA_COUNT]

        # log that we're about to create the headers
        logging.debug("About to start creating the header row.")

        # let's select cell A1 to start writing out the headers
        the_ws["A1"].value = BASIC_ITEM_HEADER
        the_ws["A1"].alignment = Alignment(horizontal='center')
        the_ws["A1"].font = Font(bold=True)
        the_ws.column_dimensions['A'].width = 9

        the_ws["B1"].value = BASIC_COLUMN_HEADER
        the_ws["B1"].font = Font(bold=True)
        the_ws.column_dimensions['B'].width = 14

        the_ws["C1"].value = BASIC_COUNT_HEADER
        the_ws["C1"].alignment = Alignment(horizontal='center')
        the_ws["C1"].font = Font(bold=True)
        the_ws.column_dimensions['C'].width = 8

        the_ws["D1"].value = NA_COLUMN_LABEL
        the_ws["D1"].alignment = Alignment(horizontal='center')
        the_ws["D1"].font = Font(bold=True)
        the_ws.column_dimensions['D'].width = 10

        the_ws["E1"].value = DISTINCT_VALUE_COUNT
        the_ws["E1"].alignment = Alignment(horizontal='center')
        the_ws["E1"].font = Font(bold=True)
        the_ws.column_dimensions['E'].width = 19

        the_ws["F1"].value = VALUE_UNIQUE_HEADER
        the_ws["F1"].alignment = Alignment(horizontal='center')
        the_ws["F1"].font = Font(bold=True)
        the_ws.column_dimensions['F'].width = 15

        the_ws["G1"].value = VALUES_ACTUAL
        the_ws["G1"].alignment = Alignment(horizontal='left')
        the_ws["G1"].font = Font(bold=True)
        the_ws.column_dimensions['G'].width = 60

        # we need to freeze the top row
        the_ws.freeze_panes = the_ws["B2"]

        # log that we're about to start looping over the objects
        logging.debug("About to start adding object items to page.")

        # define the row_count
        row_count = 1

        # loop over the items in the list
        for next_value in the_list:
            # log the first retrieved item
            logging.debug(f"retrieved object [{next_value}]")

            # write out the item #
            the_ws["A1"].offset(row=row_count, column=0).value = count_dict[next_value]
            the_ws["A1"].offset(row=row_count, column=0).alignment = Alignment(horizontal='center')

            # write out the column name
            the_ws["A1"].offset(row=row_count, column=1).value = next_value
            the_ws["A1"].offset(row=row_count, column=1).alignment = Alignment(horizontal='left')

            # write out the count, will not include N/A
            the_ws["A1"].offset(row=row_count, column=2).value = self.data_analyzer.the_df[next_value].count()
            the_ws["A1"].offset(row=row_count, column=2).alignment = Alignment(horizontal='center')
            the_ws["A1"].offset(row=row_count, column=2).number_format = FORMAT_NUMBER_COMMA

            # write out the NaN count
            the_ws["A1"].offset(row=row_count, column=3).value = nan_count_dict[next_value]
            the_ws["A1"].offset(row=row_count, column=3).alignment = Alignment(horizontal='center')
            the_ws["A1"].offset(row=row_count, column=3).number_format = FORMAT_NUMBER_COMMA

            # We can have a None present in unique_values[next_value] for the generation of the INITIAL report
            # Thus, we need to protect against that situation

            if unique_values[next_value] is None:
                # log an error message
                logging.error(f"unique_values[{next_value}][0] is NONE.")
            elif unique_values[next_value][0] == VALUES_TOO_MANY:
                the_ws["A1"].offset(row=row_count, column=4).value = len(self.data_analyzer.the_df[next_value].unique())
            else:
                # write out the count of unique values
                the_ws["A1"].offset(row=row_count, column=4).value = len(unique_values[next_value])

            the_ws["A1"].offset(row=row_count, column=4).alignment = Alignment(horizontal='center')
            the_ws["A1"].offset(row=row_count, column=4).number_format = FORMAT_NUMBER_COMMA

            the_ws["A1"].offset(row=row_count, column=5).value = unique_column[next_value]
            the_ws["A1"].offset(row=row_count, column=5).alignment = Alignment(horizontal='center')

            # write out unique values
            the_ws["A1"].offset(row=row_count, column=6).value = str(unique_values[next_value])
            the_ws["A1"].offset(row=row_count, column=6).alignment = Alignment(horizontal='left', wrap_text=True)

            # increment row_count
            row_count = row_count + 1

    # create correlations sheet
    def __create_correlations_sheet__(self, the_model_type: str, the_index: int):
        # log that we've been called
        self.logger.info("A request to generate the correlations tab has been made.")

        # Note: It is expected that this function is only run when performing a FULL report, not INITIAL.

        # retrieve the list of all correlations
        the_corr_list = self.stat_generator.get_list_of_correlations(the_type=ALL_CORRELATIONS)

        # get the dict of plots
        plot_dict = self.plot_generator.plot_storage

        # create the basic analysis tab
        self.the_workbook.create_sheet(index=the_index, title="CORRELATIONS")

        # get the sheet references
        the_ws = self.the_workbook.worksheets[the_index]

        # set the base cell that we start all calculations
        base_cell_addr = "A1"
        the_offset = 0
        row_count = 1

        # if we're logistic or linear, create a heatmap
        if the_model_type == MT_LOGISTIC_REGRESSION or the_model_type == MT_LINEAR_REGRESSION:

            # log the image we're going to import
            self.logger.debug(f"About to insert image [{plot_dict[GENERAL][PLOT_TYPE_HEATMAP]}]")

            # create an instance of the image
            the_image = Image(plot_dict[GENERAL][PLOT_TYPE_HEATMAP])

            # Anchor the image in at F3
            the_image.anchor = the_ws[base_cell_addr].offset(row=3, column=5).coordinate

            # log where we're putting the image
            self.logger.debug(f"Image anchor is [{the_image.anchor}]")

            # add the image to the worksheet
            the_ws.add_image(the_image)

        # CREATE HEADER

        # let's select cell A1 to start writing out the headers
        the_ws["A1"].value = COLUMN_HEADER + BASIC_SINGLE_SPACE + str(1)
        the_ws["A1"].alignment = Alignment(horizontal='left')
        the_ws["A1"].font = Font(bold=True)
        the_ws.column_dimensions['A'].width = 18

        # NEXT COLUMN
        col_incr = 1
        the_ws["A1"].offset(row=0, column=col_incr).value = COLUMN_HEADER + BASIC_SINGLE_SPACE + str(2)
        the_ws["A1"].offset(row=0, column=col_incr).alignment = Alignment(horizontal='left')
        the_ws["A1"].offset(row=0, column=col_incr).font = Font(bold=True)
        the_ws.column_dimensions['B'].width = 18

        # NEXT COLUMN
        col_incr = 2
        the_ws["A1"].offset(row=0, column=col_incr).value = CORRELATION_HEADER
        the_ws["A1"].offset(row=0, column=col_incr).alignment = Alignment(horizontal='center')
        the_ws["A1"].offset(row=0, column=col_incr).font = Font(bold=True)
        the_ws.column_dimensions['C'].width = 14

        # NEXT COLUMN
        col_incr = 4
        the_ws["A1"].offset(row=0, column=col_incr).value = CORRELATION_LEVEL
        the_ws["A1"].offset(row=0, column=col_incr).alignment = Alignment(horizontal='center')
        the_ws["A1"].offset(row=0, column=col_incr).font = Font(bold=True)
        the_ws.column_dimensions['E'].width = 11

        # set the border around label for the level, and the cell that holds the significance level
        self.set_border(the_ws, CellRange("E1:F1"))

        # write the level out to cell F1
        the_ws["A1"].offset(row=0, column=5).value = self.stat_generator.get_correlation_level()
        the_ws["A1"].offset(row=0, column=5).alignment = Alignment(horizontal='center')
        the_ws["A1"].offset(row=0, column=5).font = Font(bold=True)

        # change the color of F1 to green.
        the_ws["A1"].offset(row=0, column=5).fill = PatternFill(start_color='0000FF00',
                                                                end_color='0000FF00',
                                                                fill_type='solid')

        # we need to freeze the top row.  I'm not sure why, but you have to be one row below what you think.
        the_ws.freeze_panes = the_ws["B2"]

        # Now, loop over all the correlations
        for current_corr in the_corr_list:
            # log what correlation we are working with
            self.logger.debug(f"writing out correlation {current_corr}")

            # write out the name of column 1
            the_ws["A1"].offset(row=row_count, column=0).value = current_corr[0]
            the_ws["A1"].offset(row=row_count, column=0).alignment = Alignment(horizontal='left')

            # write out the name of column 2
            the_ws["A1"].offset(row=row_count, column=1).value = current_corr[1]
            the_ws["A1"].offset(row=row_count, column=1).alignment = Alignment(horizontal='left')

            # write out the correlation
            the_ws["A1"].offset(row=row_count, column=2).value = current_corr[2]
            the_ws["A1"].offset(row=row_count, column=2).alignment = Alignment(horizontal='center')
            the_ws["A1"].offset(row=row_count, column=2).number_format = '0.00000'

            # mark a correlation as significant if it exceeds the level
            if current_corr[2] >= self.stat_generator.get_correlation_level():
                the_ws["A1"].offset(row=row_count, column=2).font = Font(bold=True)
                the_ws["A1"].offset(row=row_count, column=2).fill = PatternFill(start_color='0000FF00',
                                                                                end_color='0000FF00',
                                                                                fill_type='solid')

            # increment the row
            row_count = row_count + 1

        # next, reset the starting point for the individual plots
        row_count = 30

        # check if the scatter plots exist on the plot_generator.  We only want to display the relavant ones.
        if PLOT_TYPE_SCATTER_CHART in plot_dict:
            # log that we've started add scatter plots
            self.logger.debug("Adding scatter plots to tab.")

            # get the list of plots
            for the_current_plot in plot_dict[PLOT_TYPE_SCATTER_CHART]:
                # log the current plot
                self.logger.debug(f"attaching scatter plot [{the_current_plot}]")

                # create an instance of the image
                the_image = Image(plot_dict[PLOT_TYPE_SCATTER_CHART][the_current_plot])

                # Anchor the image on column F
                the_image.anchor = the_ws[base_cell_addr].offset(row=row_count, column=5).coordinate

                # add the image to the worksheet
                the_ws.add_image(the_image)

                # increment the row_count
                row_count = row_count + 30

    # create chi-squared sheet
    def __create_chi_squared_sheet__(self, the_index):
        # log that we've been called
        self.logger.info("A request to generate the chi-squared tab has been made.")

        # retrieve the list of chi-squared relationships
        the_chisqd_list = self.stat_generator.get_chi_squared_results()

        # create the basic analysis tab
        self.the_workbook.create_sheet(index=the_index, title="CHI-SQRD RESULTS")

        # get the sheet references
        the_ws = self.the_workbook.worksheets[the_index]

        # set the base cell that we start all calculations
        base_cell_addr = "A1"
        the_offset = 0
        row_count = 1

        # CREATE HEADER

        # let's select cell A1 to start writing out the headers
        the_ws["A1"].value = COLUMN_HEADER + BASIC_SINGLE_SPACE + str(1)
        the_ws["A1"].alignment = Alignment(horizontal='left')
        the_ws["A1"].font = Font(bold=True)
        the_ws.column_dimensions['A'].width = 18

        # NEXT COLUMN
        col_incr = 1
        the_ws["A1"].offset(row=0, column=col_incr).value = COLUMN_HEADER + BASIC_SINGLE_SPACE + str(2)
        the_ws["A1"].offset(row=0, column=col_incr).alignment = Alignment(horizontal='left')
        the_ws["A1"].offset(row=0, column=col_incr).font = Font(bold=True)
        the_ws.column_dimensions['B'].width = 18

        # NEXT COLUMN
        col_incr = 2
        the_ws["A1"].offset(row=0, column=col_incr).value = CHI_SQUARED_HEADER
        the_ws["A1"].offset(row=0, column=col_incr).alignment = Alignment(horizontal='center')
        the_ws["A1"].offset(row=0, column=col_incr).font = Font(bold=True)
        the_ws.column_dimensions['C'].width = 20

        # we need to freeze the top row.  I'm not sure why, but you have to be one row below what you think.
        the_ws.freeze_panes = the_ws["B2"]

        # Now, loop over all the chi-squared results
        for current_result in the_chisqd_list:
            # log what chi-squared result we are working with
            self.logger.debug(f"writing out chi-squared {current_result}")

            # write out the name of column 1
            the_ws["A1"].offset(row=row_count, column=0).value = current_result[0]
            the_ws["A1"].offset(row=row_count, column=0).alignment = Alignment(horizontal='left')

            # write out the name of column 2
            the_ws["A1"].offset(row=row_count, column=1).value = current_result[1]
            the_ws["A1"].offset(row=row_count, column=1).alignment = Alignment(horizontal='left')

            # write out the chi-squared result
            the_ws["A1"].offset(row=row_count, column=2).value = current_result[2]
            the_ws["A1"].offset(row=row_count, column=2).alignment = Alignment(horizontal='center')
            the_ws["A1"].offset(row=row_count, column=2).number_format = '0.00000'

            # check if the chi-squared results is < .05, then fill in a color.
            if current_result[2] < .05:
                the_ws["A1"].offset(row=row_count, column=2).font = Font(bold=True)
                the_ws["A1"].offset(row=row_count, column=2).fill = PatternFill(start_color='0000FF00',
                                                                                end_color='0000FF00',
                                                                                fill_type='solid')

            # increment the row
            row_count = row_count + 1

    # create linear model results sheet
    def __create_model_results_sheet__(self, the_index: int, the_type: str, the_model_type: str):
        # validate the_type argument
        if the_type not in ANALYSIS_TYPE:
            raise RuntimeError("unknown type passed to __create_model_results_sheet__()")
        elif the_index <= 0:
            raise AttributeError("the_index must be greater than zero.")

        # log that we've been called
        self.logger.info("A request to generate the model results tab has been made.")

        # variable declaration
        tab_title = self.get_model_type_title(the_model_type=the_model_type, the_run_type=the_type)

        # create the tab name
        self.the_workbook.create_sheet(index=the_index, title=tab_title)

        # create the tab name
        if the_type == ANALYZE_DATASET_INITIAL:
            # retrieve the initial model result
            the_model_result = self.initial_model.get_the_result()
        elif the_type == ANALYZE_DATASET_FULL and the_model_type in [MT_KNN_CLASSIFICATION, MT_RF_REGRESSION]:
            # retrieve the initial model result
            the_model_result = self.initial_model.get_the_result()
        else:
            # retrieve the final model result
            the_model_result = self.final_model.get_the_result()

        # retrieve the_results_df
        the_results_df = the_model_result.get_results_dataframe()

        # retrieve the list of features
        feature_list = the_model_result.get_the_variables_list()

        # get the sheet references
        the_ws = self.the_workbook.worksheets[the_index]

        # get the dict of plots
        plot_dict = self.plot_generator.plot_storage

        # set the base cell that we start all calculations
        base_cell_addr = "A1"
        the_offset = 0
        row_count = 1

        # CREATE REPORT HEADER TABLE

        # COLUMN A - TARGET - HEADER
        the_ws[base_cell_addr].offset(row=the_offset, column=0).value = TARGET_HEADER
        the_ws[base_cell_addr].offset(row=the_offset, column=0).alignment = Alignment(horizontal='left')
        the_ws[base_cell_addr].offset(row=the_offset, column=0).font = Font(bold=True)
        the_ws[base_cell_addr].offset(row=the_offset, column=0).border = \
            Border(top=BORDER_THICK, left=BORDER_THICK, right=BORDER_THICK, bottom=BORDER_THICK)
        the_ws.column_dimensions['A'].width = 30

        # COLUMN B - TARGET FEATURE - Actual Value
        the_ws[base_cell_addr].offset(row=the_offset, column=1).value = the_model_result.get_the_target_variable()
        the_ws[base_cell_addr].offset(row=the_offset, column=1).alignment = Alignment(horizontal='center')
        the_ws[base_cell_addr].offset(row=the_offset, column=1).font = Font(bold=True)
        the_ws[base_cell_addr].offset(row=the_offset, column=1).border = \
            Border(top=BORDER_THICK, left=BORDER_THICK, right=BORDER_THICK, bottom=BORDER_THICK)
        the_ws.column_dimensions['B'].width = 36

        # additional column width settings
        the_ws.column_dimensions['C'].width = 12
        the_ws.column_dimensions['D'].width = 16
        the_ws.column_dimensions['E'].width = 16
        the_ws.column_dimensions['F'].width = 12
        the_ws.column_dimensions['G'].width = 16
        the_ws.column_dimensions['H'].width = 16
        the_ws.column_dimensions['I'].width = 16

        # CELL A2 - NUMBER OF FEATURES HEADER
        the_offset = the_offset + 1
        the_ws[base_cell_addr].offset(row=the_offset, column=0).value = NUMBER_OF_FEATURES_HEADER
        the_ws[base_cell_addr].offset(row=the_offset, column=0).alignment = Alignment(horizontal='left')
        the_ws[base_cell_addr].offset(row=the_offset, column=0).font = Font(bold=True)

        # CELL B2 - NUMBER OF FEATURES - Actual Value
        the_ws[base_cell_addr].offset(row=the_offset, column=1).value = len(the_model_result.get_the_variables_list())
        the_ws[base_cell_addr].offset(row=the_offset, column=1).alignment = Alignment(horizontal='center')

        # CELL A3 - CONSTANT HEADER
        if the_model_result.has_assumption(the_assumption=MODEL_CONSTANT):
            the_offset = the_offset + 1
            the_ws[base_cell_addr].offset(row=the_offset, column=0).value = CONSTANT_HEADER
            the_ws[base_cell_addr].offset(row=the_offset, column=0).alignment = Alignment(horizontal='left')
            the_ws[base_cell_addr].offset(row=the_offset, column=0).font = Font(bold=True)

            the_ws[base_cell_addr].offset(row=the_offset, column=1).value = \
                the_model_result.get_assumption_result(the_assumption=MODEL_CONSTANT)
            the_ws[base_cell_addr].offset(row=the_offset, column=1).alignment = Alignment(horizontal='center')

        # PSEUDO_R_SQUARED_HEADER
        if the_model_result.has_assumption(the_assumption=PSEUDO_R_SQUARED_HEADER):
            the_offset = the_offset + 1
            the_ws[base_cell_addr].offset(row=the_offset, column=0).value = PSEUDO_R_SQUARED_HEADER
            the_ws[base_cell_addr].offset(row=the_offset, column=0).alignment = Alignment(horizontal='left')
            the_ws[base_cell_addr].offset(row=the_offset, column=0).font = Font(bold=True)

            the_ws[base_cell_addr].offset(row=the_offset, column=1).value = \
                the_model_result.get_assumption_result(the_assumption=PSEUDO_R_SQUARED_HEADER)
            the_ws[base_cell_addr].offset(row=the_offset, column=1).alignment = Alignment(horizontal='center')
            the_ws[base_cell_addr].offset(row=the_offset, column=1).number_format = FORMAT_PERCENTAGE

        # R_SQUARED_HEADER
        if the_model_result.has_assumption(the_assumption=R_SQUARED_HEADER):
            the_offset = the_offset + 1
            the_ws[base_cell_addr].offset(row=the_offset, column=0).value = R_SQUARED_HEADER
            the_ws[base_cell_addr].offset(row=the_offset, column=0).alignment = Alignment(horizontal='left')
            the_ws[base_cell_addr].offset(row=the_offset, column=0).font = Font(bold=True)

            the_ws[base_cell_addr].offset(row=the_offset, column=1).value = \
                the_model_result.get_assumption_result(the_assumption=R_SQUARED_HEADER)
            the_ws[base_cell_addr].offset(row=the_offset, column=1).alignment = Alignment(horizontal='center')
            the_ws[base_cell_addr].offset(row=the_offset, column=1).number_format = FORMAT_PERCENTAGE

        # ADJ_R_SQUARED_HEADER
        if the_model_result.has_assumption(the_assumption=ADJ_R_SQUARED_HEADER):
            the_offset = the_offset + 1
            the_ws[base_cell_addr].offset(row=the_offset, column=0).value = ADJ_R_SQUARED_HEADER
            the_ws[base_cell_addr].offset(row=the_offset, column=0).alignment = Alignment(horizontal='left')
            the_ws[base_cell_addr].offset(row=the_offset, column=0).font = Font(bold=True)

            the_ws[base_cell_addr].offset(row=the_offset, column=1).value = \
                the_model_result.get_assumption_result(the_assumption=ADJ_R_SQUARED_HEADER)
            the_ws[base_cell_addr].offset(row=the_offset, column=1).alignment = Alignment(horizontal='center')
            the_ws[base_cell_addr].offset(row=the_offset, column=1).number_format = FORMAT_PERCENTAGE

        # CELL A6 - AIC
        if the_model_result.has_assumption(the_assumption=AIC_SCORE):
            the_offset = the_offset + 1
            the_ws[base_cell_addr].offset(row=the_offset, column=0).value = AIC_HEADER
            the_ws[base_cell_addr].offset(row=the_offset, column=0).alignment = Alignment(horizontal='left')
            the_ws[base_cell_addr].offset(row=the_offset, column=0).font = Font(bold=True)

            the_ws[base_cell_addr].offset(row=the_offset, column=1).value = \
                the_model_result.get_assumption_result(the_assumption=AIC_SCORE)
            the_ws[base_cell_addr].offset(row=the_offset, column=1).alignment = Alignment(horizontal='center')

        # CELL A7 - BIC
        if the_model_result.has_assumption(the_assumption=BIC_SCORE):
            the_offset = the_offset + 1
            the_ws[base_cell_addr].offset(row=the_offset, column=0).value = BIC_HEADER
            the_ws[base_cell_addr].offset(row=the_offset, column=0).alignment = Alignment(horizontal='left')
            the_ws[base_cell_addr].offset(row=the_offset, column=0).font = Font(bold=True)

            the_ws[base_cell_addr].offset(row=the_offset, column=1).value = \
                the_model_result.get_assumption_result(the_assumption=BIC_SCORE)
            the_ws[base_cell_addr].offset(row=the_offset, column=1).alignment = Alignment(horizontal='center')

        # CELL A8 - F-STATISTIC
        if the_model_result.has_assumption(the_assumption=F_STATISTIC_HEADER):
            the_offset = the_offset + 1
            the_ws[base_cell_addr].offset(row=the_offset, column=0).value = F_STATISTIC_HEADER
            the_ws[base_cell_addr].offset(row=the_offset, column=0).alignment = Alignment(horizontal='left')
            the_ws[base_cell_addr].offset(row=the_offset, column=0).font = Font(bold=True)

            the_ws[base_cell_addr].offset(row=the_offset, column=1).value = \
                the_model_result.get_assumption_result(the_assumption=F_STATISTIC_HEADER)
            the_ws[base_cell_addr].offset(row=the_offset, column=1).alignment = Alignment(horizontal='center')

        # CELL A9 - P VALUE OF F-STATISTIC
        if the_model_result.has_assumption(the_assumption=P_VALUE_F_STATISTIC_HEADER):
            the_offset = the_offset + 1
            the_ws[base_cell_addr].offset(row=the_offset, column=0).value = P_VALUE_F_STATISTIC_HEADER
            the_ws[base_cell_addr].offset(row=the_offset, column=0).alignment = Alignment(horizontal='left')
            the_ws[base_cell_addr].offset(row=the_offset, column=0).font = Font(bold=True)

            the_ws[base_cell_addr].offset(row=the_offset, column=1).value = \
                the_model_result.get_assumption_result(the_assumption=P_VALUE_F_STATISTIC_HEADER)
            the_ws[base_cell_addr].offset(row=the_offset, column=1).alignment = Alignment(horizontal='center')

        # CELL A10 - LOG LIKELIHOOD
        if the_model_result.has_assumption(the_assumption=LOG_LIKELIHOOD):
            the_offset = the_offset + 1
            the_ws[base_cell_addr].offset(row=the_offset, column=0).value = LOG_LIKELIHOOD
            the_ws[base_cell_addr].offset(row=the_offset, column=0).alignment = Alignment(horizontal='left')
            the_ws[base_cell_addr].offset(row=the_offset, column=0).font = Font(bold=True)

            the_ws[base_cell_addr].offset(row=the_offset, column=1).value = \
                the_model_result.get_assumption_result(the_assumption=LOG_LIKELIHOOD)
            the_ws[base_cell_addr].offset(row=the_offset, column=1).alignment = Alignment(horizontal='center')

        # CELL A11 - NUMBER OF OBS
        if the_model_result.has_assumption(the_assumption=NUMBER_OF_OBS):
            the_offset = the_offset + 1
            the_ws[base_cell_addr].offset(row=the_offset, column=0).value = NUMBER_OF_OBS
            the_ws[base_cell_addr].offset(row=the_offset, column=0).alignment = Alignment(horizontal='left')
            the_ws[base_cell_addr].offset(row=the_offset, column=0).font = Font(bold=True)

            the_ws[base_cell_addr].offset(row=the_offset, column=1).value = \
                the_model_result.get_assumption_result(the_assumption=NUMBER_OF_OBS)
            the_ws[base_cell_addr].offset(row=the_offset, column=1).alignment = Alignment(horizontal='center')

        # CELL A12 - DEGREES OF FREEDOM (model)
        if the_model_result.has_assumption(the_assumption=DEGREES_OF_FREEDOM_MODEL):
            the_offset = the_offset + 1
            the_ws[base_cell_addr].offset(row=the_offset, column=0).value = DEGREES_OF_FREEDOM_MODEL
            the_ws[base_cell_addr].offset(row=the_offset, column=0).alignment = Alignment(horizontal='left')
            the_ws[base_cell_addr].offset(row=the_offset, column=0).font = Font(bold=True)

            the_ws[base_cell_addr].offset(row=the_offset, column=1).value = \
                the_model_result.get_assumption_result(the_assumption=DEGREES_OF_FREEDOM_MODEL)
            the_ws[base_cell_addr].offset(row=the_offset, column=1).alignment = Alignment(horizontal='center')

        # CELL A13 - DEGREES OF FREEDOM (residuals)
        if the_model_result.has_assumption(the_assumption=DEGREES_OF_FREEDOM_RESID):
            the_offset = the_offset + 1
            the_ws[base_cell_addr].offset(row=the_offset, column=0).value = DEGREES_OF_FREEDOM_RESID
            the_ws[base_cell_addr].offset(row=the_offset, column=0).alignment = Alignment(horizontal='left')
            the_ws[base_cell_addr].offset(row=the_offset, column=0).font = Font(bold=True)

            the_ws[base_cell_addr].offset(row=the_offset, column=1).value = \
                the_model_result.get_assumption_result(the_assumption=DEGREES_OF_FREEDOM_RESID)
            the_ws[base_cell_addr].offset(row=the_offset, column=1).alignment = Alignment(horizontal='center')

        # DURBAN WATSON STATISTIC
        if the_model_result.has_assumption(the_assumption=DURBAN_WATSON_STATISTIC):
            the_offset = the_offset + 1
            the_ws[base_cell_addr].offset(row=the_offset, column=0).value = DURBAN_WATSON_STATISTIC
            the_ws[base_cell_addr].offset(row=the_offset, column=0).alignment = Alignment(horizontal='left')
            the_ws[base_cell_addr].offset(row=the_offset, column=0).font = Font(bold=True)

            the_ws[base_cell_addr].offset(row=the_offset, column=1).value = \
                the_model_result.get_assumption_result(the_assumption=DURBAN_WATSON_STATISTIC)
            the_ws[base_cell_addr].offset(row=the_offset, column=1).alignment = Alignment(horizontal='center')

        # RESIDUAL STANDARD ERROR
        if the_model_result.has_assumption(the_assumption=RESIDUAL_STD_ERROR):
            the_offset = the_offset + 1
            the_ws[base_cell_addr].offset(row=the_offset, column=0).value = RESIDUAL_STD_ERROR
            the_ws[base_cell_addr].offset(row=the_offset, column=0).alignment = Alignment(horizontal='left')
            the_ws[base_cell_addr].offset(row=the_offset, column=0).font = Font(bold=True)

            the_ws[base_cell_addr].offset(row=the_offset, column=1).value = \
                the_model_result.get_assumption_result(the_assumption=RESIDUAL_STD_ERROR)
            the_ws[base_cell_addr].offset(row=the_offset, column=1).alignment = Alignment(horizontal='center')

        # MODEL MEAN ABSOLUTE ERROR
        if the_model_result.has_assumption(the_assumption=MODEL_MEAN_ABSOLUTE_ERROR):
            the_offset = the_offset + 1
            the_ws[base_cell_addr].offset(row=the_offset, column=0).value = MODEL_MEAN_ABSOLUTE_ERROR
            the_ws[base_cell_addr].offset(row=the_offset, column=0).alignment = Alignment(horizontal='left')
            the_ws[base_cell_addr].offset(row=the_offset, column=0).font = Font(bold=True)

            the_ws[base_cell_addr].offset(row=the_offset, column=1).value = \
                the_model_result.get_assumption_result(the_assumption=MODEL_MEAN_ABSOLUTE_ERROR)
            the_ws[base_cell_addr].offset(row=the_offset, column=1).alignment = Alignment(horizontal='center')

        # MODEL MEAN SQUARED ERROR
        if the_model_result.has_assumption(the_assumption=MODEL_MEAN_SQUARED_ERROR):
            the_offset = the_offset + 1
            the_ws[base_cell_addr].offset(row=the_offset, column=0).value = MODEL_MEAN_SQUARED_ERROR
            the_ws[base_cell_addr].offset(row=the_offset, column=0).alignment = Alignment(horizontal='left')
            the_ws[base_cell_addr].offset(row=the_offset, column=0).font = Font(bold=True)

            the_ws[base_cell_addr].offset(row=the_offset, column=1).value = \
                the_model_result.get_assumption_result(the_assumption=MODEL_MEAN_SQUARED_ERROR)
            the_ws[base_cell_addr].offset(row=the_offset, column=1).alignment = Alignment(horizontal='center')

        # MODEL ROOT MEAN SQUARED ERROR
        if the_model_result.has_assumption(the_assumption=MODEL_ROOT_MEAN_SQUARED_ERROR):
            the_offset = the_offset + 1
            the_ws[base_cell_addr].offset(row=the_offset, column=0).value = MODEL_ROOT_MEAN_SQUARED_ERROR
            the_ws[base_cell_addr].offset(row=the_offset, column=0).alignment = Alignment(horizontal='left')
            the_ws[base_cell_addr].offset(row=the_offset, column=0).font = Font(bold=True)

            the_ws[base_cell_addr].offset(row=the_offset, column=1).value = \
                the_model_result.get_assumption_result(the_assumption=MODEL_ROOT_MEAN_SQUARED_ERROR)
            the_ws[base_cell_addr].offset(row=the_offset, column=1).alignment = Alignment(horizontal='center')

        # BREUSCH PAGAN TESTS
        if the_model_result.has_assumption(the_assumption=BREUSCH_PAGAN_P_VALUE):
            the_offset = the_offset + 1
            the_ws[base_cell_addr].offset(row=the_offset, column=0).value = LM_LAGRANGE_MULTIPLIER_STATISTIC
            the_ws[base_cell_addr].offset(row=the_offset, column=0).alignment = Alignment(horizontal='left')
            the_ws[base_cell_addr].offset(row=the_offset, column=0).font = Font(bold=True)

            the_ws[base_cell_addr].offset(row=the_offset, column=1).value = \
                the_model_result.get_assumption_result(BREUSCH_PAGAN_P_VALUE)[LM_LAGRANGE_MULTIPLIER_STATISTIC]
            the_ws[base_cell_addr].offset(row=the_offset, column=1).alignment = Alignment(horizontal='center')

            the_offset = the_offset + 1
            the_ws[base_cell_addr].offset(row=the_offset, column=0).value = BREUSCH_PAGAN_P_VALUE
            the_ws[base_cell_addr].offset(row=the_offset, column=0).alignment = Alignment(horizontal='left')
            the_ws[base_cell_addr].offset(row=the_offset, column=0).font = Font(bold=True)

            the_ws[base_cell_addr].offset(row=the_offset, column=1).value = \
                the_model_result.get_assumption_result(BREUSCH_PAGAN_P_VALUE)[LM_P_VALUE]
            the_ws[base_cell_addr].offset(row=the_offset, column=1).alignment = Alignment(horizontal='center')

        # JARQUE-BERA TESTS
        if the_model_result.has_assumption(the_assumption=JARQUE_BERA_STATISTIC):
            the_offset = the_offset + 1
            the_ws[base_cell_addr].offset(row=the_offset, column=0).value = LM_JARQUE_BERA_STATISTIC
            the_ws[base_cell_addr].offset(row=the_offset, column=0).alignment = Alignment(horizontal='left')
            the_ws[base_cell_addr].offset(row=the_offset, column=0).font = Font(bold=True)

            the_ws[base_cell_addr].offset(row=the_offset, column=1).value = \
                the_model_result.get_assumption_result(JARQUE_BERA_STATISTIC)[LM_JARQUE_BERA_STATISTIC]
            the_ws[base_cell_addr].offset(row=the_offset, column=1).alignment = Alignment(horizontal='center')

            # JARQUE-BERA TEST - PROBABILITY
            the_offset = the_offset + 1
            the_ws[base_cell_addr].offset(row=the_offset, column=0).value = LM_JARQUE_BERA_PROB
            the_ws[base_cell_addr].offset(row=the_offset, column=0).alignment = Alignment(horizontal='left')
            the_ws[base_cell_addr].offset(row=the_offset, column=0).font = Font(bold=True)

            the_ws[base_cell_addr].offset(row=the_offset, column=1).value = \
                the_model_result.get_assumption_result(JARQUE_BERA_STATISTIC)[LM_JARQUE_BERA_PROB]
            the_ws[base_cell_addr].offset(row=the_offset, column=1).alignment = Alignment(horizontal='center')

        # CREATE FEATURE LISTING HEADER
        # retrieve the list of columns present in the results dataframe
        feature_column_list = the_model_result.get_feature_columns()

        # offset two rows
        the_offset = the_offset + 2

        # define column offset
        column_offset = 0

        # COLUMN A - FEATURE_NUMBER
        if LM_FEATURE_NUM in feature_column_list:
            the_offset = the_offset + 1
            the_ws[base_cell_addr].offset(row=the_offset, column=column_offset).value = FEATURE_NUMBER
            the_ws[base_cell_addr].offset(row=the_offset, column=column_offset).alignment = \
                Alignment(horizontal='center')
            the_ws[base_cell_addr].offset(row=the_offset, column=column_offset).font = Font(bold=True)

        # COLUMN B - FEATURE_NAME_HEADER
        column_offset = 1
        the_ws[base_cell_addr].offset(row=the_offset, column=column_offset).value = FEATURE_NAME_HEADER
        the_ws[base_cell_addr].offset(row=the_offset, column=column_offset).alignment = Alignment(horizontal='left')
        the_ws[base_cell_addr].offset(row=the_offset, column=column_offset).font = Font(bold=True)

        # COLUMN C - COEFFICIENT_HEADER
        if LM_COEFFICIENT in feature_column_list:
            column_offset = column_offset + 1
            the_ws[base_cell_addr].offset(row=the_offset, column=column_offset).value = COEFFICIENT_HEADER
            the_ws[base_cell_addr].offset(row=the_offset, column=column_offset).alignment = \
                Alignment(horizontal='center')
            the_ws[base_cell_addr].offset(row=the_offset, column=column_offset).font = Font(bold=True)

        # COLUMN D - STANDARD_ERROR
        if LM_STANDARD_ERROR in feature_column_list:
            column_offset = column_offset + 1
            the_ws[base_cell_addr].offset(row=the_offset, column=column_offset).value = STANDARD_ERROR_HEADER
            the_ws[base_cell_addr].offset(row=the_offset, column=column_offset).alignment = \
                Alignment(horizontal='center')
            the_ws[base_cell_addr].offset(row=the_offset, column=column_offset).font = Font(bold=True)

        # COLUMN E - T_STATISTIC_HEADER
        if LM_T_STATISTIC in feature_column_list:
            column_offset = column_offset + 1
            the_ws[base_cell_addr].offset(row=the_offset, column=column_offset).value = T_STATISTIC_HEADER
            the_ws[base_cell_addr].offset(row=the_offset, column=column_offset).alignment = \
                Alignment(horizontal='center')
            the_ws[base_cell_addr].offset(row=the_offset, column=column_offset).font = Font(bold=True)

        # COLUMN F - P_VALUE_HEADER
        if LM_P_VALUE in feature_column_list:
            column_offset = column_offset + 1
            the_ws[base_cell_addr].offset(row=the_offset, column=column_offset).value = P_VALUE_HEADER
            the_ws[base_cell_addr].offset(row=the_offset, column=column_offset).alignment = \
                Alignment(horizontal='center')
            the_ws[base_cell_addr].offset(row=the_offset, column=column_offset).font = Font(bold=True)

        # COLUMN G - '[0.025'
        if LM_LS_CONF_INT in feature_column_list:
            column_offset = column_offset + 1
            the_ws[base_cell_addr].offset(row=the_offset, column=column_offset).value = LM_LS_CONF_INT
            the_ws[base_cell_addr].offset(row=the_offset, column=column_offset).alignment = \
                Alignment(horizontal='center')
            the_ws[base_cell_addr].offset(row=the_offset, column=column_offset).font = Font(bold=True)

        # COLUMN H - '0.975]'
        if LM_RS_CONF_INT in feature_column_list:
            column_offset = column_offset + 1
            the_ws[base_cell_addr].offset(row=the_offset, column=column_offset).value = LM_RS_CONF_INT
            the_ws[base_cell_addr].offset(row=the_offset, column=column_offset).alignment = \
                Alignment(horizontal='center')
            the_ws[base_cell_addr].offset(row=the_offset, column=column_offset).font = Font(bold=True)

        # COLUMN I - VIF
        if LM_VIF in feature_column_list:
            column_offset = column_offset + 1
            the_ws[base_cell_addr].offset(row=the_offset, column=column_offset).value = VIF_HEADER
            the_ws[base_cell_addr].offset(row=the_offset, column=column_offset).alignment = \
                Alignment(horizontal='center')
            the_ws[base_cell_addr].offset(row=the_offset, column=column_offset).font = Font(bold=True)

        # loop over the feature_list
        for the_feature in feature_list:
            # increment the offset
            the_offset = the_offset + 1
            the_column = 0

            # COLUMN A - LM_FEATURE_NUM - ACTUAL VALUE
            the_ws[base_cell_addr].offset(row=the_offset, column=the_column).value = \
                the_results_df[LM_FEATURE_NUM].loc[the_feature]
            the_ws[base_cell_addr].offset(row=the_offset, column=the_column).alignment = Alignment(horizontal='center')

            # COLUMN B - FEATURE_NAME_HEADER - ACTUAL VALUE
            the_column = the_column + 1
            the_ws[base_cell_addr].offset(row=the_offset, column=the_column).value = the_feature
            the_ws[base_cell_addr].offset(row=the_offset, column=the_column).alignment = Alignment(horizontal='left')

            # COLUMN C - LM_COEFFICIENT - ACTUAL VALUE
            if LM_COEFFICIENT in feature_column_list:
                the_column = the_column + 1
                the_ws[base_cell_addr].offset(row=the_offset, column=the_column).value = \
                    the_results_df[LM_COEFFICIENT].loc[the_feature]
                the_ws[base_cell_addr].offset(row=the_offset, column=the_column).alignment = \
                    Alignment(horizontal='center')

            # COLUMN D - LM_STANDARD_ERROR - ACTUAL VALUE
            if LM_STANDARD_ERROR in feature_column_list:
                the_column = the_column + 1
                the_ws[base_cell_addr].offset(row=the_offset, column=the_column).value = \
                    the_results_df[LM_STANDARD_ERROR].loc[the_feature]
                the_ws[base_cell_addr].offset(row=the_offset, column=the_column).alignment = \
                    Alignment(horizontal='center')

            # COLUMN E - LM_T_STATISTIC - ACTUAL VALUE
            if LM_T_STATISTIC in feature_column_list:
                the_column = the_column + 1
                the_ws[base_cell_addr].offset(row=the_offset, column=the_column).value = \
                    the_results_df[LM_T_STATISTIC].loc[the_feature]
                the_ws[base_cell_addr].offset(row=the_offset, column=the_column).alignment = \
                    Alignment(horizontal='center')

            # COLUMN F - LM_P_VALUE - ACTUAL VALUE
            if LM_P_VALUE in feature_column_list:
                the_column = the_column + 1
                the_ws[base_cell_addr].offset(row=the_offset, column=the_column).value = \
                    the_results_df[LM_P_VALUE].loc[the_feature]
                the_ws[base_cell_addr].offset(row=the_offset, column=the_column).alignment = \
                    Alignment(horizontal='center')

            # COLUMN G - '[0.025' or LM_LS_CONF_INT
            if LM_LS_CONF_INT in feature_column_list:
                the_column = the_column + 1
                the_ws[base_cell_addr].offset(row=the_offset, column=the_column).value = \
                    the_results_df[LM_LS_CONF_INT].loc[the_feature]
                the_ws[base_cell_addr].offset(row=the_offset, column=the_column).alignment = \
                    Alignment(horizontal='center')

            # COLUMN H - '0.975]' or LM_RS_CONF_INT
            if LM_RS_CONF_INT in feature_column_list:
                the_column = the_column + 1
                the_ws[base_cell_addr].offset(row=the_offset, column=the_column).value = \
                    the_results_df[LM_RS_CONF_INT].loc[the_feature]
                the_ws[base_cell_addr].offset(row=the_offset, column=the_column).alignment = \
                    Alignment(horizontal='center')

            # COLUMN I - LM_VIF - ACTUAL VALUE
            if LM_VIF in feature_column_list:
                the_column = the_column + 1
                the_ws[base_cell_addr].offset(row=the_offset, column=the_column).value = \
                    the_results_df[LM_VIF].loc[the_feature]
                the_ws[base_cell_addr].offset(row=the_offset, column=the_column).alignment = \
                    Alignment(horizontal='center')

        # increment the_offset by 3
        the_offset = the_offset + 3
        assumption_count = 0

        # determine if this is a full analysis, and we have a model that handles MODEL ACCURACY
        if the_type == ANALYZE_DATASET_FULL and the_model_result.has_assumption(MODEL_ACCURACY):
            the_ws[base_cell_addr].offset(row=the_offset, column=0).value = MODEL_ACCURACY
            the_ws[base_cell_addr].offset(row=the_offset, column=0).alignment = Alignment(horizontal='left')
            the_ws[base_cell_addr].offset(row=the_offset, column=0).font = Font(bold=True)

            the_ws[base_cell_addr].offset(row=the_offset, column=1).value = \
                the_model_result.get_assumption_result(MODEL_ACCURACY)
            the_ws[base_cell_addr].offset(row=the_offset, column=1).alignment = Alignment(horizontal='center')
            the_ws[base_cell_addr].offset(row=the_offset, column=1).number_format = FORMAT_PERCENTAGE

            # increment the offset and assumption_count
            the_offset = the_offset + 1
            assumption_count = assumption_count + 1

        # determine if this is a full analysis, and we have a model that handles MODEL PRECISION
        if the_type == ANALYZE_DATASET_FULL and the_model_result.has_assumption(MODEL_PRECISION):
            the_ws[base_cell_addr].offset(row=the_offset, column=0).value = MODEL_PRECISION
            the_ws[base_cell_addr].offset(row=the_offset, column=0).alignment = Alignment(horizontal='left')
            the_ws[base_cell_addr].offset(row=the_offset, column=0).font = Font(bold=True)

            the_ws[base_cell_addr].offset(row=the_offset, column=1).value = \
                the_model_result.get_assumption_result(MODEL_PRECISION)
            the_ws[base_cell_addr].offset(row=the_offset, column=1).alignment = Alignment(horizontal='center')
            the_ws[base_cell_addr].offset(row=the_offset, column=1).number_format = FORMAT_PERCENTAGE

            # increment the offset and assumption_count
            the_offset = the_offset + 1
            assumption_count = assumption_count + 1

        # determine if this is a full analysis, and we have a model that handles MODEL AVG PRECISION
        if the_type == ANALYZE_DATASET_FULL and the_model_result.has_assumption(MODEL_AVG_PRECISION):
            the_ws[base_cell_addr].offset(row=the_offset, column=0).value = MODEL_AVG_PRECISION
            the_ws[base_cell_addr].offset(row=the_offset, column=0).alignment = Alignment(horizontal='left')
            the_ws[base_cell_addr].offset(row=the_offset, column=0).font = Font(bold=True)

            the_ws[base_cell_addr].offset(row=the_offset, column=1).value = \
                the_model_result.get_assumption_result(MODEL_AVG_PRECISION)
            the_ws[base_cell_addr].offset(row=the_offset, column=1).alignment = Alignment(horizontal='center')
            the_ws[base_cell_addr].offset(row=the_offset, column=1).number_format = FORMAT_PERCENTAGE

            # increment the offset and assumption_count
            the_offset = the_offset + 1
            assumption_count = assumption_count + 1

        # determine if this is a full analysis, and we have a model that handles MODEL RECALL
        if the_type == ANALYZE_DATASET_FULL and the_model_result.has_assumption(MODEL_RECALL):
            the_ws[base_cell_addr].offset(row=the_offset, column=0).value = MODEL_RECALL
            the_ws[base_cell_addr].offset(row=the_offset, column=0).alignment = Alignment(horizontal='left')
            the_ws[base_cell_addr].offset(row=the_offset, column=0).font = Font(bold=True)

            the_ws[base_cell_addr].offset(row=the_offset, column=1).value = \
                    the_model_result.get_assumption_result(MODEL_RECALL)
            the_ws[base_cell_addr].offset(row=the_offset, column=1).alignment = Alignment(horizontal='center')
            the_ws[base_cell_addr].offset(row=the_offset, column=1).number_format = FORMAT_PERCENTAGE

            # increment the offset and assumption_count
            the_offset = the_offset + 1
            assumption_count = assumption_count + 1

        # determine if this is a full analysis, and we have a model that handles MODEL F1 SCORE
        if the_type == ANALYZE_DATASET_FULL and the_model_result.has_assumption(MODEL_F1_SCORE):
            the_ws[base_cell_addr].offset(row=the_offset, column=0).value = MODEL_F1_SCORE
            the_ws[base_cell_addr].offset(row=the_offset, column=0).alignment = Alignment(horizontal='left')
            the_ws[base_cell_addr].offset(row=the_offset, column=0).font = Font(bold=True)

            the_ws[base_cell_addr].offset(row=the_offset, column=1).value = \
                the_model_result.get_assumption_result(MODEL_F1_SCORE)
            the_ws[base_cell_addr].offset(row=the_offset, column=1).alignment = Alignment(horizontal='center')
            the_ws[base_cell_addr].offset(row=the_offset, column=1).number_format = FORMAT_PERCENTAGE

            # increment the offset and assumption_count
            the_offset = the_offset + 1
            assumption_count = assumption_count + 1

        # determine if this is a full analysis, and we have a model that handles MODEL ROC SCORE
        if the_type == ANALYZE_DATASET_FULL and the_model_result.has_assumption(MODEL_ROC_SCORE):
            the_ws[base_cell_addr].offset(row=the_offset, column=0).value = MODEL_ROC_SCORE
            the_ws[base_cell_addr].offset(row=the_offset, column=0).alignment = Alignment(horizontal='left')
            the_ws[base_cell_addr].offset(row=the_offset, column=0).font = Font(bold=True)

            the_ws[base_cell_addr].offset(row=the_offset, column=1).value = \
                the_model_result.get_assumption_result(MODEL_ROC_SCORE)
            the_ws[base_cell_addr].offset(row=the_offset, column=1).alignment = Alignment(horizontal='center')
            the_ws[base_cell_addr].offset(row=the_offset, column=1).number_format = FORMAT_PERCENTAGE

            # increment the offset and assumption_count
            the_offset = the_offset + 1
            assumption_count = assumption_count + 1

        # determine if this is a full analysis, and we have a model that handles MODEL BEST SCORE
        if the_type == ANALYZE_DATASET_FULL and the_model_result.has_assumption(MODEL_BEST_SCORE):
            the_ws[base_cell_addr].offset(row=the_offset, column=0).value = MODEL_BEST_SCORE
            the_ws[base_cell_addr].offset(row=the_offset, column=0).alignment = Alignment(horizontal='left')
            the_ws[base_cell_addr].offset(row=the_offset, column=0).font = Font(bold=True)

            the_ws[base_cell_addr].offset(row=the_offset, column=1).value = \
                the_model_result.get_assumption_result(MODEL_BEST_SCORE)
            the_ws[base_cell_addr].offset(row=the_offset, column=1).alignment = Alignment(horizontal='center')

            # increment the offset and assumption_count
            the_offset = the_offset + 1
            assumption_count = assumption_count + 1

        # determine if this is a full analysis, and we have a model that handles MODEL BEST PARAMS
        if the_type == ANALYZE_DATASET_FULL and the_model_result.has_assumption(MODEL_BEST_PARAMS):
            the_ws[base_cell_addr].offset(row=the_offset, column=0).value = MODEL_BEST_PARAMS
            the_ws[base_cell_addr].offset(row=the_offset, column=0).alignment = Alignment(horizontal='left')
            the_ws[base_cell_addr].offset(row=the_offset, column=0).font = Font(bold=True)

            the_ws[base_cell_addr].offset(row=the_offset, column=1).value = \
                the_model_result.get_assumption_result(MODEL_BEST_PARAMS)
            the_ws[base_cell_addr].offset(row=the_offset, column=1).alignment = Alignment(horizontal='left')

            # increment the offset and assumption_count
            the_offset = the_offset + 1
            assumption_count = assumption_count + 1

        # If this is a KNN or logistic regression model, there will be a confusion matrix created
        # we can detect this by checking for the existence of the key PLOT_TYPE_CM_HEATMAP in plot_dict[GENERAL]
        if GENERAL in plot_dict:
            # we need to reduce the offset by assumption_count.  I need to rethink this logic in the future
            the_offset = the_offset - assumption_count

            # add confusion matrix heatmap if it was generated for the model.
            if PLOT_TYPE_CM_HEATMAP in plot_dict[GENERAL]:
                # create an instance of the confusion matrix heat map
                the_image = Image(plot_dict[GENERAL][PLOT_TYPE_CM_HEATMAP])
                the_image.anchor = the_ws[base_cell_addr].offset(row=the_offset, column=4).coordinate

                # log where we're putting the image
                logging.info(f"Image anchor is [{the_image.anchor}]")

                # add the image to the worksheet
                the_ws.add_image(the_image)

                # increment the_offset by 26
                the_offset = the_offset + 26

            # add ROC/AUC plot if it was generated for the model.
            if PLOT_TYPE_ROC_AUC in plot_dict[GENERAL]:
                # create an instance of the confusion matrix heat map
                the_image = Image(plot_dict[GENERAL][PLOT_TYPE_ROC_AUC])
                the_image.anchor = the_ws[base_cell_addr].offset(row=the_offset, column=4).coordinate

                # log where we're putting the image
                logging.info(f"Image anchor is [{the_image.anchor}]")

                # add the image to the worksheet
                the_ws.add_image(the_image)

                # increment the_offset by 26
                the_offset = the_offset + 26

        # check if we are a linear or logistical model
        if the_model_type == MT_LINEAR_REGRESSION or the_model_type == MT_LOGISTIC_REGRESSION:
            # now we need to inject the residual plots into the Excel file
            for the_feature in feature_list:
                # log the feature we are about to add the residual plot for
                self.logger.debug(f"Adding residual plot for [{the_feature}]")

                # COLUMN A - FEATURE - HEADER
                the_ws[base_cell_addr].offset(row=the_offset, column=0).value = FEATURE_NAME_HEADER
                the_ws[base_cell_addr].offset(row=the_offset, column=0).alignment = Alignment(horizontal='left')
                the_ws[base_cell_addr].offset(row=the_offset, column=0).font = Font(bold=True)
                the_ws[base_cell_addr].offset(row=the_offset, column=0).border = \
                    Border(top=BORDER_THICK, left=BORDER_THICK, right=BORDER_THICK, bottom=BORDER_THICK)

                # COLUMN B - FEATURE - Actual Value
                the_ws[base_cell_addr].offset(row=the_offset, column=1).value = the_feature
                the_ws[base_cell_addr].offset(row=the_offset, column=1).alignment = Alignment(horizontal='center')
                the_ws[base_cell_addr].offset(row=the_offset, column=1).font = Font(bold=True)
                the_ws[base_cell_addr].offset(row=the_offset, column=1).border = \
                    Border(top=BORDER_THICK, left=BORDER_THICK, right=BORDER_THICK, bottom=BORDER_THICK)

                # jump to the next row, and leave a two row gap.
                the_offset = the_offset + 3

                # ***************************************************************************
                #                                 RESIDUAL PLOTS
                # ***************************************************************************

                # check if the plots are available
                if the_feature in plot_dict:
                    if PLOT_TYPE_STD_RESIDUAL in plot_dict[the_feature]:
                        # create an instance of the standardized residual plot image
                        the_image = Image(plot_dict[the_feature][PLOT_TYPE_STD_RESIDUAL])
                        the_image.anchor = the_ws[base_cell_addr].offset(row=the_offset, column=1).coordinate

                        # log where we're putting the image
                        logging.info(f"Image anchor is [{the_image.anchor}]")

                        # add the image to the worksheet
                        the_ws.add_image(the_image)

                # ***************************************************************************
                #                                 LONG ODDS PLOTS
                # ***************************************************************************

                if PLOT_TYPE_LONG_ODDS in plot_dict:
                    # create an instance of the long odds plot image
                    the_image = Image(plot_dict[PLOT_TYPE_LONG_ODDS][the_feature])
                    the_image.anchor = the_ws[base_cell_addr].offset(row=the_offset, column=1).coordinate

                    # log where we're putting the image
                    logging.info(f"Image anchor is [{the_image.anchor}]")

                    # add the image to the worksheet
                    the_ws.add_image(the_image)
                else:
                    # log a message
                    self.logger.debug(f"feature[{the_feature}] was not present.")

                # ***************************************************************************
                #                                 QQ RESIDUAL PLOTS
                # ***************************************************************************

                if PLOT_TYPE_Q_Q_RESIDUAL_PLOT in plot_dict:
                    # create an instance of the QQ residual plot image
                    the_image = Image(plot_dict[PLOT_TYPE_Q_Q_RESIDUAL_PLOT][the_feature])
                    the_image.anchor = the_ws[base_cell_addr].offset(row=the_offset, column=7).coordinate

                    # log where we're putting the image
                    logging.info(f"Image anchor is [{the_image.anchor}]")

                    # add the image to the worksheet
                    the_ws.add_image(the_image)

                # jump several rows
                the_offset = the_offset + 26

    # set a border around a range
    def set_border(self, ws, cell_range):
        logging.debug(f"setting border on [{str(ws)}][{cell_range}]")
        rows = ws[cell_range.coord]
        side = Side(border_style='thick', color="00000000")

        rows = list(rows)  # we convert iterator to list for simplicity, but it's not memory efficient solution
        max_y = len(rows) - 1  # index of the last row
        for pos_y, cells in enumerate(rows):
            max_x = len(cells) - 1  # index of the last cell
            for pos_x, cell in enumerate(cells):
                border = Border(
                    left=cell.border.left,
                    right=cell.border.right,
                    top=cell.border.top,
                    bottom=cell.border.bottom
                )
                if pos_x == 0:
                    border.left = side
                if pos_x == max_x:
                    border.right = side
                if pos_y == 0:
                    border.top = side
                if pos_y == max_y:
                    border.bottom = side

                # set new border only if it's one of the edge cells
                if pos_x == 0 or pos_x == max_x or pos_y == 0 or pos_y == max_y:
                    cell.border = border

    # export dataframes to excel
    def export_dataframes_to_excel(self, dataset_analyzer, pca_analysis):
        # run validations
        if not self.is_valid(dataset_analyzer, DatasetAnalyzer):
            raise SyntaxError("The dataset analyzer was None or incorrect type.")
        elif not self.is_valid(dataset_analyzer.the_df, DataFrame):
            raise SyntaxError("Cannot export dataframe since it has not been created.")
        elif not self.is_valid(dataset_analyzer.the_normal_df, DataFrame):
            raise SyntaxError("Cannot export normalized dataframe since it has not been created.")
        elif not self.is_valid(pca_analysis, PCA_Analysis):
            raise SyntaxError("The PCA_Analysis was None or incorrect type.")

        # log that we've been called
        logging.debug("A request to export internal dataframes to Excel.")

        # instantiate an ExcelManager
        excel_manager = ExcelManager()

        # log that we're creating output of dataframes
        logging.debug(f"generating output file [{self.dataframe_path}]")

        # create the workbook
        excel_manager.create_workbook(the_path=self.dataframe_path)

        # open the workbook
        the_wb = excel_manager.open_workbook(the_path=self.dataframe_path)

        # export the cleaned dataframe
        excel_manager.write_df_into_wb_tab(dataset_analyzer.the_original_df, REPORT_ORIG_DF_TAB_NAME, the_wb)

        # export the cleaned dataframe
        excel_manager.write_df_into_wb_tab(dataset_analyzer.the_df, REPORT_CLEAN_DF_TAB_NAME, the_wb)

        # export the normalized dataframe
        excel_manager.write_df_into_wb_tab(dataset_analyzer.the_normal_df, REPORT_NORM_DF_TAB_NAME, the_wb)

        # export the PCA loadings per component
        excel_manager.write_df_into_wb_tab(pca_analysis.pca_loadings, REPORT_PAC_TAB_NAME, the_wb, display_index=True)

        # close workbook
        excel_manager.close_workbook(the_path=self.dataframe_path)

        # log that we've been called
        logging.debug("Export of dataframe completed.")

    # get the model type title for tab.
    def get_model_type_title(self, the_model_type: str, the_run_type: str) -> str:
        # run validations
        if the_model_type not in MT_OPTIONS:
            raise SyntaxError("the_model_type is None or incorrect option.")
        elif the_run_type not in ANALYSIS_TYPE:
            raise SyntaxError("the_run_type is None or incorrect option.")

        # variable declarations
        the_result = None

        # check if we have a linear regression model
        if the_model_type == MT_LINEAR_REGRESSION:
            # determine if it's initial or final
            if the_run_type == ANALYZE_DATASET_INITIAL:
                the_result = "INITIAL LINEAR MODEL RESULTS"
            else:
                the_result = "FINAL LINEAR MODEL RESULTS"
        # check if we have a logistic regression model
        elif the_model_type == MT_LOGISTIC_REGRESSION:
            # determine if it's initial or final
            if the_run_type == ANALYZE_DATASET_INITIAL:
                the_result = "INITIAL LOGISTIC MODEL RESULTS"
            else:
                the_result = "FINAL LOGISTIC MODEL RESULTS"
        # check if we have a KNN
        elif the_model_type == MT_KNN_CLASSIFICATION:
            the_result = "KNN MODEL RESULT"
        # check if we have a Random Forest
        elif the_model_type == MT_RF_REGRESSION:
            the_result = "RANDOM FOREST MODEL RESULT"

        # return
        return the_result
