import logging
import numpy as np
import pandas as pd
import openpyxl

from openpyxl.workbook import Workbook
from os.path import exists
from pandas import DataFrame, Timestamp
from openpyxl.utils.dataframe import dataframe_to_rows
from model.constants.ExcelConstants import EMPTY_STRING, ILLEGAL_TAB_FIELDS, SINGLE_SPACE, LETTER_KEY
from util.BaseUtil import BaseUtil
from util.CommonUtils import strip_preceding_letter_from_field, strip_trailing_digit_from_field


def as_text(value):
    if value is None:
        return ""
    return str(value)


# class holding utilities for working with Excel documents.
class ExcelManager(BaseUtil):
    # constants
    NA_VALUE_LIST = ['#N/A', '<NA>', 'NA', np.nan, np.NaN, np.NAN, pd.NA]
    EMPTY_STRING = ""

    # init method
    def __init__(self):
        super().__init__()

        # define logger
        self.logger = logging.getLogger(__name__)

        # log that we've been instantiated
        self.logger.debug("A request to create an ExcelManager has been made.")

        # internal storage
        self.wb_storage = {}

    # create an Excel Workbook
    def create_workbook(self, the_path):
        # perform validation
        if not self.is_valid(the_path, str):
            raise SyntaxError("The path was None or incorrect type.")

        # log that we've been called
        logging.debug(f"creating a Workbook at [{the_path}]")

        # create a workbook
        the_wb = openpyxl.Workbook()

        # save the workbook to file system
        the_wb.save(the_path)

    # open an Excel Workbook
    def open_workbook(self, the_path):
        # perform validation
        if not self.is_valid(the_path, str):
            raise SyntaxError("The path was None or incorrect type.")
        elif not exists(the_path):
            raise FileNotFoundError(f"The path [{the_path}] does not exist.")

        # log that we've been called
        logging.debug(f"A request to open [{the_path}] has been made.")

        # open the workbook
        the_result = openpyxl.load_workbook(the_path)

        # add to storage
        self.wb_storage[the_path] = the_result

        # return the workbook
        return the_result

    # close and save a workbook
    def close_workbook(self, the_path):
        # perform validation
        if not self.is_valid(the_path, str):
            raise SyntaxError("The path was None or incorrect type.")
        elif not exists(the_path):
            raise FileNotFoundError(f"The path [{the_path}] does not exist.")
        elif not the_path in self.wb_storage:
            raise KeyError(f"The path [{the_path}] was not in storage.")

        # log that we've been called
        logging.debug(f"A request to close workbook at [{the_path}] has been made.")

        # get the workbook from storage
        the_wb = self.wb_storage[the_path]

        # save workbook
        the_wb.save(the_path)

        # close workbook
        the_wb.close()

        # remove from storage
        del self.wb_storage[the_path]

    # write the contents of a dataframe into a workbook
    def write_df_into_wb_tab(self, the_df, tab_name, the_wb, display_index=False):
        # perform validation
        if not self.is_valid(the_df, DataFrame):
            raise SyntaxError("The incoming DataFrame was None or incorrect type.")
        elif not self.is_valid(tab_name, str):
            raise SyntaxError("The incoming tab name was None or incorrect type.")
        elif not self.is_valid(the_wb, Workbook):
            raise SyntaxError("The incoming workbook was None or incorrect type.")

        # log that we've been called
        logging.debug(f"Generating tab[{tab_name} for DataFrame.")

        # variable declaration
        the_ws = None

        # check if the tab exists on the workbook
        if tab_name in the_wb.sheetnames:
            # log that the worksheet existed
            logging.debug(f"overwriting worksheet[{tab_name}]")

            # get a reference to the worksheet
            the_ws = the_wb.get_sheet_by_name(tab_name)

            # remove the existing workbook
            the_wb.remove(the_ws)

        # log that we are creating the worksheet
        logging.debug(f"creating worksheet[{tab_name}]")

        # create an empty sheet [sheet_name] using old index
        the_wb.create_sheet(tab_name)

        # get a reference to the worksheet
        the_ws = the_wb[tab_name]

        # get a row objects
        rows = dataframe_to_rows(the_df, index=display_index)

        # loop over all rows
        for r_idx, row in enumerate(rows, 1):
            # loop over all columns
            for c_idx, value in enumerate(row, 1):
                # check if value is NA
                if value is pd.NA:
                    # write out blank.
                    the_ws.cell(row=r_idx, column=c_idx, value=EMPTY_STRING)
                # we shouldn't get tuples, but if we do, handle it.
                elif isinstance(value, tuple):
                    # write out the value in the first element.
                    the_ws.cell(row=r_idx, column=c_idx, value=value[0])
                elif type(value) is Timestamp:
                    the_ws.cell(row=r_idx, column=c_idx, value=value)
                    the_ws.number_format = "mm/dd/yyyy;@"
                # this is what we want to do
                else:
                    # write out the value
                    the_ws.cell(row=r_idx, column=c_idx, value=value)

        # auto-size the columns
        self.autosize_worksheet(the_ws)

        # log that we are complete
        logging.debug(f"finished writing out contents to workbook[{tab_name}]")

    # auto_size a worksheet
    @staticmethod
    def autosize_worksheet(the_ws):
        for column_cells in the_ws.columns:
            length = max(len(as_text(cell.value)) for cell in column_cells)

            # log the length for the column
            logging.debug(f"the length will be set to {length + 4}")

            # change the column width
            the_ws.column_dimensions[openpyxl.utils.get_column_letter(column_cells[0].column)].width = length + 4

    # get clean tab name.
    def get_clean_tab_name(self, name_to_use):
        # run validation
        if not self.is_valid(name_to_use, str):
            raise ValueError(f"the name_to_use argument was None or incorrect type.")

        # log that we've been called
        self.logger.debug(f"cleaning tab name[{name_to_use}].")

        # variable declaration
        the_result = name_to_use

        # loop through all characters
        for current_char in range(0, len(name_to_use)):
            if name_to_use[current_char] in ILLEGAL_TAB_FIELDS:
                the_result = name_to_use.replace(name_to_use[current_char], SINGLE_SPACE)

        # return
        return the_result

    # increment column.
    # Note, this needs to be re-implemented using base 26 addition.  I didn't think about this solution until
    # I was already finished with the below.
    def increment_column(self, the_col, amount):
        # define the result
        the_result = None

        # reverse the column
        temp_col = the_col[::-1]

        # convert the first letter to a value
        column_value = LETTER_KEY[temp_col[0]]

        # check if the sum is > 26
        if column_value + amount > 26:
            # get the quotient and modulus
            the_quotient = int((column_value + amount) / 26)
            the_modulus = (column_value + amount) % 26

            # if the quotient is 1, and there is only one letter
            if the_quotient == 1 and len(temp_col) == 1:
                # we need to increment the first letter of the reversed string by the modulus
                first_letter = self.increment_column(temp_col[0], the_modulus)

                # set the result
                the_result = first_letter + "A"
            elif the_quotient == 1 and len(temp_col) == 2:
                # set the result
                the_result = self.increment_column(temp_col[0],
                                                   the_modulus) + self.increment_column(temp_col[1:2], the_quotient)
            else:
                # we need to increment the rest of the letter by the quotient

                # set the result
                the_result = self.increment_column(temp_col[0],
                                                   the_modulus) + self.increment_column(temp_col[1:2],
                                                                                        the_quotient) + temp_col[2:]

        # we're staying within the alphabet
        else:
            # get the letter associated with adjustment
            new_letter = list(LETTER_KEY.keys())[list(LETTER_KEY.values()).index(column_value + amount)]

            # now, we have to check if we only have one character or multiple
            if len(the_col) == 1:
                the_result = new_letter
            else:
                the_result = new_letter + temp_col[1:]

        # return the final result
        return the_result[::-1]

    # offset an Excel address cell in string form
    def offset_excel_address(self, original_address, col_offset, row_offset):
        # variable definition
        the_letter = None
        the_number = None
        the_result = None

        # log that we've been called
        logging.debug("A request to offset_excel_address() has been made. ")

        # run validations
        if not self.is_valid(original_address, str):
            raise TypeError("the original address is None or incorrect type.")
        elif not self.is_valid(col_offset, int):
            raise TypeError("the column offset is None or incorrect type.")
        elif not self.is_valid(row_offset, int):
            raise TypeError("the row offset is None or incorrect type.")

        # log the incoming arguments
        logging.debug("original_address[%s], col_offset[%s], row_offset[%s]",
                      original_address, col_offset, row_offset)

        # strip the letter
        the_letter = strip_preceding_letter_from_field(original_address).upper()

        # strip the number
        the_number = strip_trailing_digit_from_field(original_address)

        # log the stripped letter and number
        logging.debug("The original letter[%s], the original number[%s]", the_letter, the_number)

        # offset the letter
        the_letter = self.increment_column(the_letter, col_offset)

        # offset the number
        if the_number + row_offset >= 1:
            the_number = the_number + row_offset
        else:
            raise SyntaxError("the row value cannot be less than 1.")

        # assemble the result
        the_result = the_letter + str(the_number)

        # log the final result
        logging.debug("The final result is [%s]", the_result)

        # return statement
        return the_result
